import os
import re
import sys
import json
import xlrd
import random
import chardet
import hashlib
import threading
import openpyxl

from tqdm import tqdm
from pathlib import Path
from textgrid import TextGrid
from pydub import AudioSegment
from collections import defaultdict

class InputCollectBase:
    def __init__(self, data_dir, kformat_dir, cut2short=False, mode=None, test_rate=0.001, lang='zh-CN'):
        """
        Initialize.
        Args:
            data_dir: input data directory
            kformat_dir: directory for output kaldi-format files
            cut2short: cut audios to short or not
            mode: train or None
        """
        self.data_dir = data_dir
        self.kformat_dir = kformat_dir
        self.cut2short = cut2short
        self.mode = mode
        self.test_rate = test_rate
        self.lang = lang

    def ReadMD5(self, filenames):
        print(f"Starting calculate the md5 value...")
        for ii, filename in enumerate(filenames):
            if ii % 100 == 0:
                print(f"MD5 | [{ii}/{len(filenames)}] {filename}...")
            if os.path.exists(filename+".md5"):
                continue
            cons = open(filename, 'rb').read()
            file_md5 = hashlib.md5(cons).hexdigest()
            with open(filename+".md5", 'w', encoding='utf-8') as f:
                f.write(f"{file_md5}\n")
        print("DONE")


    def ReadMD5Thread(self, wav_files, npart=30):
        random.shuffle(wav_files)
        nums_perpart = len(wav_files) // npart + 1
        parts = [wav_files[l:l+nums_perpart] for l in range(0, len(wav_files), nums_perpart)]
        print(len(wav_files))
        threads = []
        for part in parts:
            t = threading.Thread(target=self.ReadMD5, args=(part,))
            threads.append(t)
            t.start()
        for t in threads:
            t.join()

    def readDuration(self, wav_file):
        dur_file = wav_file + ".dur"
        if os.path.exists(dur_file):
            try:
                dur = float(open(dur_file).read().strip())
            except Exception:
                sound = AudioSegment.from_wav(wav_file)
                dur = sound.duration_seconds
                with open(dur_file, 'w', encoding='utf-8') as f:
                    f.write(f"{dur}\n")
        else:
            sound = AudioSegment.from_wav(wav_file)
            dur = sound.duration_seconds
            # with open(dur_file, 'w', encoding='utf-8') as f:
            #     f.write(f"{dur}\n")
        return dur

    def writeDuration(self, wav_files):
        print(f"Starting calculate the md5 value...")
        for ii, wav_file in enumerate(wav_files):
            if ii % 1000 == 0:
                print(f"DUR | [{ii}/{len(wav_files)}] {wav_file}...")
            dur_file = wav_file + ".dur"
            if os.path.exists(dur_file):
                continue
            try:
                sound = AudioSegment.from_wav(wav_file)
            except Exception as e:
                print(f"ERROR!!! [{wav_file}] with {e}")
            dur = sound.duration_seconds
            with open(dur_file, 'w', encoding='utf-8') as f:
                f.write(f"{dur}\n")

    def writeDurationThread(self, wav_files, npart=30):
        random.shuffle(wav_files)
        nums_perpart = len(wav_files) // npart + 1
        parts = [wav_files[l:l+nums_perpart] for l in range(0, len(wav_files), nums_perpart)]
        threads = []
        for part in parts:
            t = threading.Thread(target=self.writeDuration, args=(part,))
            threads.append(t)
            t.start()
        for t in threads:
            t.join()

    def check(self, text):
        """
        Check if there are invalid characters.
        """
        # if len(re.sub(r"[a-zA-Z]+", "", text))==0:
        #     return text
        text = text.replace(" ", "")
        if text in ["muisc", "music", "tts", "noi", "sil"]:
            return text
        if self.lang == "zh-CN":
            no_char = re.sub(r"[\u4e00-\u9fa5]", "", text) # Remove all chinese characters
        elif self.lang == "ko-KR":
            no_char = re.sub(r"[\uac00-\ud7ff]", "", text) # Remove all chinese characters
        elif self.lang == "ja-JP":
            no_char = re.sub(r"[\u4E00-\u9FFF\u3040-\u309f\u30a0-\u30ff]", "", text) # Remove all chinese characters
            no_char = re.sub("[グ〆〒﨑グ〇々]", "", no_char)
        else:
            no_char = ""
        no_char = re.sub(r"[a-zA-Z]+", "", no_char)
        no_char = re.sub(f"'", "", no_char)
        no_char = re.sub("[，。？！.?:!,]", '', no_char).replace(" ", "")
        return no_char

    def strQ2B(self, ustring):
        """全角转半角"""
        rstring = ""
        for uchar in ustring:
            inside_code=ord(uchar)
            if inside_code == 12288:                              #全角空格直接转换            
                inside_code = 32 
            elif (inside_code >= 65281 and inside_code <= 65374): #全角字符（除空格）根据关系转化
                inside_code -= 65248
            rstring += chr(inside_code)
        return rstring

    def process(self):
        """
        Generate kaldi-format files.
        """
        raise NotImplementedError

    def find(self):
        """
        Find all wav files or text files.
        """
        pass

    def findOne(self, dir_, format_):
        """
        Using pathlib.Path.rglob to find files. and transfer the pathObj type to list type
        Args:
            dir_: the directory for files -> str
            format_: the format for the files which need to be found -> str
        Return:
            obj_file: list of objective format files. -> list
        """
        
        obj_files = Path(dir_).rglob(format_)
        obj_files = [str(ele) for ele in obj_files]
        return obj_files

    @property
    def info(self):
        """
        Get the information of current data. Such as:
        1. #wave and #text (if there is)
        2. original path of data
        3. kaldi-format data dict
        4. ...
        """
        raise NotImplementedError

    def kformatFiles(self, kdir):
        """
        Return the path of wav.scp
        """
        wavscp_file = os.path.join(kdir, 'wav.scp')
        u2s_file = os.path.join(kdir, 'utt2spk')
        s2u_file = os.path.join(kdir, 'spk2utt')
        transcript_file = os.path.join(kdir, 'text')
        segments_file = os.path.join(kdir, 'segments')
        utt2dur_file = os.path.join(kdir, 'utt2dur')
        return wavscp_file, transcript_file, u2s_file, s2u_file, segments_file, utt2dur_file

    def cut2ShortWave(self):
        """
        if cut2short is True, the long audio will cutted to short audios.
        and the short audios will be saved into self.kformat_dir + "/short"
        """
        pass

    def write2Files(self, kdir, name2wavpath, name2text, name2spk, names, long=False):
        wavscp_file, transcript_file, u2s_file, s2u_file, segments_file, utt2dur_file = self.kformatFiles(kdir)
        if long:
            spk2name = defaultdict(list)
            print(f"=>> Number of Names: {len(names)}")
            with open(wavscp_file, 'w', encoding='utf-8') as fw, \
                open(transcript_file, 'w', encoding='utf-8') as ft, \
                open(u2s_file, 'w', encoding='utf-8') as fu, \
                open(s2u_file, 'w', encoding='utf-8') as fs, \
                open(segments_file, 'w', encoding='utf-8') as fseg, \
                open(utt2dur_file, 'w', encoding='utf-8') as fdur:
                for name in names:
                    if name not in name2wavpath.keys():
                        print("WAVE NOT FOUND!: ", name)
                        continue
                    if " " in name2wavpath[name]:
                        continue
                    if name not in name2text.keys():
                        print(f"{name} not in name2text...")
                        continue
                    try:
                        fw.write(f"{name} {name2wavpath[name]}\n")
                    except UnicodeEncodeError:
                        continue
                    for subtpr in name2text[name]:
                        text, *other_info, start_time, end_time = subtpr
                        text = text.upper()
                        if len(text) == 0:
                            print(f"Text Length == 0?")
                            continue
                        if other_info:
                            subname = f"{name}--{other_info}--{start_time}--{end_time}"
                        else:
                            subname = f"{name}--{start_time}--{end_time}"
                        fseg.write(f"{subname} {name} {start_time} {end_time}\n")
                        ft.write(f"{subname} {text}\n")
                        fu.write(f"{subname} {name2spk[name]}\n")
                        fdur.write(f"{subname} {end_time - start_time}\n")
                        spk2name[name2spk[name]].append(subname)

                for spk, subnames in spk2name.items():
                    fs.write(f"{spk}")
                    for subname in subnames:
                        fs.write(f" {subname}")
                    fs.write("\n")
        else:
            spk2name = defaultdict(list)
            with open(wavscp_file, 'w', encoding='utf-8') as fw, \
                open(transcript_file, 'w', encoding='utf-8') as ft, \
                open(u2s_file, 'w', encoding='utf-8') as fu, \
                open(s2u_file, 'w', encoding='utf-8') as fs:
                for name in names:
                    fw.write(f"{name} {name2wavpath[name]}\n")
                    ft.write(f"{name} {name2text[name].upper()}\n")
                    fu.write(f"{name} {name2spk[name]}\n")
                    spk2name[name2spk[name]].append(name)

                for spk, subnames in spk2name.items():
                    fs.write(f"{spk}")
                    for subname in subnames:
                        fs.write(f" {subname}")
                    fs.write("\n")

    def readEncodingType(self, path):
        """
        Get the encoding type of transcribed files.
        """
        content = open(path, 'rb').read()
        encoding_format = chardet.detect(content)['encoding']
        return encoding_format

    def splitAndWrite(self, spk2name, name2spk, name2text, name2wavpath):
        train_dir = os.path.join(self.kformat_dir, 'train')
        os.makedirs(train_dir, exist_ok=True)
        if self.mode == "train":
            print("=> Split data into training and testing data...")
            train_name_file, test_name_file = os.path.join(self.kformat_dir, 'train.lst'), os.path.join(self.kformat_dir, 'test.lst')
            if os.path.exists(train_name_file) and os.path.exists(test_name_file) \
                and  os.path.getsize(train_name_file) and os.path.getsize(test_name_file):
                train_names = [ele.strip() for ele in open(train_name_file).readlines()]
                test_names = [ele.strip() for ele in open(test_name_file).readlines()]
            else:
                train_names, test_names = self.splitTrainAndTest(list(spk2name.keys()), spk2name)
                with open(train_name_file, 'w', encoding='utf-8') as ftr, \
                    open(test_name_file, 'w', encoding='utf-8') as fte:
                    for ele in train_names:
                        ftr.write(f"{ele}\n")
                    for ele in test_names:
                        fte.write(f"{ele}\n")
            print(f"# Train {len(train_names)} # Test {len(test_names)}...")
            test_dir = os.path.join(self.kformat_dir, 'test')
            os.makedirs(test_dir, exist_ok=True)
            self.write2Files(train_dir, name2wavpath, name2text, name2spk, train_names, long=True)
            self.write2Files(test_dir, name2wavpath, name2text, name2spk, test_names, long=True)
        else:
            self.write2Files(train_dir, name2wavpath, name2text, name2spk, list(name2spk.keys()), long=True)


    def cleanText(self, text):
        """
        Clean text to remove invailable tags or punctuations if needed!
        """
        text = self.strQ2B(text)
        text = re.sub("【.*?】", "", text)
        text = re.sub("\[.*?\]", '', text)
        text = re.sub('<.*?>', ' ', text)
        text = re.sub('\(.*?\)', ' ', text)
        text = re.sub('[，。？,；、:`！──－：‘’“”~《》．"；,/×—－─（）・／—"『〈〉』·「」<>\.\?\!…;\-……]', ' ', text)
        text = re.sub("--'([\u4e00-\u9fa5]+)'", lambda x:x.groups()[0], text)
        text = text.replace('*', ' ')
        text = " ".join(text.split())
        if "{" in text: # 有误读的 大括号里面写的还是正常的句子，这种读错的我们不要
            text = ""   # 例如 我 看 真正 的 安定团结 是 肯定 的 ， 现在 是 说 日 <B>{渐gai4}</B> 巩固
        if self.lang == "zh-CN":
            pie = re.sub(r"[\u4e00-\u9fa5]", "", text).replace(" ", "")
            if len(list(set(pie))) == 1 and pie[0] == "'":
                print(text)
                text = text.replace("'", " ")
        text = re.compile(r" +").sub(" ", text)
        text = re.compile(r"([\u4e00-\u9fff]) +(?=[\u4e00-\u9fff])").sub("\\1", text)
        return text

    def splitTrainAndTest(self, spks, spk2name):
        if len(spks) == 1:
            wav_names = spk2name[spks[0]]
            num_test = max(int(len(wav_names) * self.test_rate), 1)

            test_names = random.sample(wav_names, num_test)
            train_names = [ele for ele in spks if ele not in wav_names]
        else:
            num_test = max(int(len(spks) * self.test_rate), 1)
            print(num_test)
            test_spks = random.sample(spks, num_test)
            train_spks = [ele for ele in spks if ele not in test_spks]
            train_names, test_names = [], []
            for spk in train_spks:
                train_names.extend(spk2name[spk])
        
            for spk in test_spks:
                test_names.extend(spk2name[spk])

        return train_names, test_names

class InputCircleIntellegenceFormat(InputCollectBase):
    """
    循环智能的数据
    Example: audio_with_lable.json  name.wav
    {"audio_filepath": "5d048930d0d06d002b674f9d_1129980_1134680.wav", "text": "衰咗啦噉無端端系得返兩條煙嗻嘛噉咪其實算失手喎，", "duration": 4.7}
    """
    def __init__(self, data_dir, kformat_dir, mode="test", test_rate=0.001):
        self.data_dir = data_dir
        self.kformat_dir = kformat_dir
        self.mode = mode
        self.test_rate = test_rate
        self.projname = os.path.basename(data_dir)
        self.info_dict = {}


    def readJsonFile(self, json_files):
        name2text = defaultdict(list)
        spk2name = defaultdict(list)
        name2spk = {}
        for json_file in json_files:
            encoding_format = self.readEncodingType(json_file)
            contents = open(json_file, encoding=encoding_format).readlines()
            for line in contents[1:]:
                line = line.strip()
                info = json.loads(line)
                name, text, duration = info["audio_filepath"], info["text"], info["duration"]
                name = self.projname + "_" + name[:-4]
                text = self.cleanText(text)
                if self.check(text):
                    print(text)
                    continue
                name2text[name].append([text, 0, duration])
                name2spk[name] = name
                spk2name[name].append(name)
        return name2text, spk2name, name2spk

    def process(self):
        json_files = self.findOne(self.data_dir, "*.json")
        wav_files = self.findOne(self.data_dir, "*.wav")

        name2text, spk2name, name2spk = self.readJsonFile(json_files)

        name2wavpath = {}
        # self.ReadMD5Thread(wav_files)
        for wav_file in tqdm(wav_files):
            filename = os.path.basename(wav_file)[:-4]
            filename = self.projname + "_" + filename
            name2wavpath[filename] = wav_file
        
        self.splitAndWrite(spk2name, name2spk, name2text, name2wavpath)

    @property
    def info(self):
        return self.info_dict

class InputAliTransFormat(InputCollectBase):
    """
    For TR Department data from alibaba.
    one result.txt containing all wave's transcripts each ID
    example: ID0000/*.wav ID0000/result.txt
    """
    def __init__(self, data_dir, kformat_dir, mode="train", test_rate=0.001, lang="zh-CN"):
        self.data_dir = data_dir
        self.kformat_dir = kformat_dir
        self.mode = mode
        self.test_rate = test_rate
        self.projname = os.path.basename(data_dir)
        self.info_dict = {}
        self.lang = lang

    def readTransFile(self, result_files):
        name2text = defaultdict(list)
        spk2name = defaultdict(list)
        name2spk = {}
        for result_file in result_files:
            encoding_format = self.readEncodingType(result_file)
            contents = open(result_file, encoding=encoding_format).readlines()
            for line in tqdm(contents[1:]):
                line = line.strip()
                try:
                    name, text, *_, time_info = line.strip().split('\t')
                except Exception:
                    print("Line Error: ", line)
                    continue
                try:
                    start_time, end_time = re.findall("\[([0-9]+\.?(?:[0-9]+)?)\]", time_info)
                except Exception:
                    print("Time Error: ", time_info)
                    continue
                start_time, end_time = float(start_time), float(end_time)
                name = self.projname + "_" + name[:-4]
                text = self.cleanText(text)
                if self.projname in ["TR2019114", "TR2019127"]:
                    text = text.replace("_", " ")
                if self.check(text):
                    print(text)
                    continue
                name2text[name] = [[text, start_time, end_time]]
                name2spk[name] = name
                spk2name[name].append(name)
        return name2text, spk2name, name2spk

    def process(self):
        result_files = self.findOne(self.data_dir, "result.txt")
        wav_files = self.findOne(self.data_dir, "*.wav")

        name2text, spk2name, name2spk = self.readTransFile(result_files)

        name2wavpath = {}
        # self.ReadMD5Thread(wav_files)
        for wav_file in tqdm(wav_files):
            filename = os.path.basename(wav_file)[:-4]
            filename = self.projname + "_" + filename
            name2wavpath[filename] = wav_file
        
        self.splitAndWrite(spk2name, name2spk, name2text, name2wavpath)


    @property
    def info(self):
        return self.info_dict

class InputDialogCorpusFormat(InputCollectBase):
    """
    For SpeechOcean Dialog Corpus
    example: A.wav A.TextGrid
    """
    def __init__(self, data_dir, kformat_dir, mode="train", test_rate=0.001, item_index=0, diff_index=None, lang="zh-CN"):
        self.data_dir = os.path.join(data_dir, "DATA")
        self.kformat_dir = kformat_dir
        self.mode = mode
        self.test_rate = test_rate
        self.item_index = item_index
        self.projname = os.path.basename(data_dir)
        self.diff_index = diff_index
        self.info_dict = {}
        self.lang = lang

    def readTextGridFile(self, tg_files):
        name2text = defaultdict(list)
        spk2name = defaultdict(list)
        name2spk = {}
        for tg_file in tg_files:
            if self.projname in ["King-ASR-447", "King-ASR-446"]:
                filename = self.projname + "_" + os.path.basename(tg_file)[:-9] + "-C0"
            elif "King-ASR-413" in tg_file:
                filename = self.projname + "_" + os.path.basename(tg_file)[:-9].split("__")[-1]
            else:
                filename = self.projname + "_" + os.path.basename(tg_file)[:-9]
            tg = TextGrid()
            try:
                tg.read(tg_file)
            except Exception as e:
                print(f"Error file {tg_file}")
                continue
            if self.diff_index is not None:
                name_num = int(os.path.basename(tg_file)[:-9])
                if name_num >= 2000: # Only for King-ASR-113 
                    self.item_index = 1
            info = tg.tiers[self.item_index]
            for ele in info:
                text, maxt, mint = ele.mark, float(ele.maxTime), float(ele.minTime)
                text = self.cleanText(text)
                if len(text) == 0:
                    continue
                if self.check(text):
                    continue
                text = text.lower().replace("#", "")
                text_with_lang = f"<{self.lang}> {text}"
                name2text[filename].append([text_with_lang, mint, maxt])
            name2spk[filename] = filename
            spk2name[filename].append(filename)
        return name2text, spk2name, name2spk

    def readTextFile(self, text_files):
        name2text = defaultdict(list)
        spk2name = defaultdict(list)
        name2spk = {}
        for text_file in text_files:
            if self.projname == "King-ASR-022":
                dirname = os.path.basename(os.path.dirname(text_file))
                filename = self.projname + "_" + dirname + "-" + os.path.basename(text_file)[:-4]
            else:
                filename = self.projname + "_" + os.path.basename(text_file)[:-4]
            encoding_format = self.readEncodingType(text_file)
            contents = open(text_file, encoding=encoding_format).readlines()
            for ele in contents:
                try:
                    subindex, timeinfo, text = ele.strip().split("\t")
                except ValueError:
                    continue
                text = text.replace("[", "").replace("]", "").replace("＋＋", "加加")
                mint, maxt = re.findall("\[([0-9]+)\-([0-9]+)\]", timeinfo)[0]
                mint, maxt = int(mint) / 1000, int(maxt) / 1000
                text = self.cleanText(text)
                if self.check(text):
                    print(text)
                    continue
                if len(text) == 0:
                    continue
                name2text[filename].append([text, mint, maxt])
            name2spk[filename] = filename
            spk2name[filename].append(filename)
        return name2text, spk2name, name2spk


    def process(self):
        if self.projname == "King-ASR-022":
            text_files = self.findOne(self.data_dir, "*.txt")
            name2text, spk2name, name2spk = self.readTextFile(text_files)
        else:
            text_files = self.findOne(self.data_dir, "*.TextGrid")
            name2text, spk2name, name2spk = self.readTextGridFile(text_files)
        wav_files = self.findOne(self.data_dir, "*.wav")
        print(f"Number of wave files: {len(text_files)}....")
        print(f"Number of wave files: {len(wav_files)}....")
        self.info_dict["num_textgrid_files"] = len(text_files)
        self.info_dict["num_wave_files"] = len(wav_files)

        name2wavpath = {}
        # # self.ReadMD5Thread(wav_files)
        for wav_file in wav_files:
            if self.projname in ["King-ASR-022"]:
                dirname = os.path.basename(os.path.dirname(wav_file))
                filename = self.projname + "_" + dirname + "-" + os.path.basename(wav_file)[:-4]
            elif self.projname in ["King-ASR-222"]:
                dirname = os.path.basename(os.path.dirname(wav_file))
                filename = self.projname + "_" + dirname + "__" + os.path.basename(wav_file)[:-4]
            elif self.projname in ["King-ASR-219"]:
                dirname = os.path.basename(os.path.dirname(wav_file))
                filename = self.projname + "_" + dirname + "_" + os.path.basename(wav_file)[:-4]
            elif self.projname in ['King-ASR-882']:
                filename = self.projname + "_"  + "_".join(os.path.basename(wav_file).split("_")[:4])
            elif self.projname in ["King-ASR-236"]:
                wav_name = os.path.basename(wav_file)
                if wav_name != "all.wav":
                    continue
                dirname = os.path.basename(os.path.dirname(wav_file))
                filename = self.projname + "_" + dirname
            else:
                filename = self.projname + "_" + os.path.basename(wav_file)[:-4]
            name2wavpath[filename] = wav_file
        
        self.splitAndWrite(spk2name, name2spk, name2text, name2wavpath)

    @property
    def info(self):
        return self.info_dict

class InputReadCorpusFormat(InputCollectBase):
    """
    For SpeechOcean Reading Corpus.
    Example:
        spk.txt   spk0001.wav spk0002.wav
    """
    def __init__(self, data_dir, kformat_dir, mode="train", test_rate=0.001, item_index=1, lang='zh-CN'):
        self.root_dir = data_dir
        self.projname = os.path.basename(data_dir)
        self.data_dir = os.path.join(data_dir, "DATA")
        self.kformat_dir = kformat_dir
        self.mode = mode
        self.item_index = item_index
        self.test_rate = test_rate
        self.lang = lang

    def readTextFile(self, txt_files, name2seg):
        name2text = {}
        spk2name = defaultdict(list)
        name2spk = {}
        for txt_file in tqdm(txt_files):
            filename = self.projname + "_" + os.path.basename(txt_file)[:-4]
            encoding_format = self.readEncodingType(txt_file)
            contents = open(txt_file, encoding=encoding_format).readlines()
            for index in range(0, len(contents), 2):
                try:
                    name, _ = contents[index].strip().split("\t")
                except ValueError:
                    name = contents[index].strip().split("\t")[0]
                name = self.projname + "_" + name
                if self.item_index == 1:
                    text = contents[index+1].strip().replace("\t", "")
                else:
                    text = contents[index].strip().split("\t")[1]
                text = self.cleanText(text)
                if self.check(text):
                    print("UncleanText: ", text)
                    continue
                text = text.lower()
                try:
                    start_time, end_time = name2seg[name]
                except Exception:
                    print(contents[index])
                    print(txt_file)
                    sys.exit(1)
                text_with_lang = f"<{self.lang}> {text}"
                name2text[name] = [[text_with_lang, start_time, end_time]]
                spk2name[filename].append(name)
                name2spk[name] = filename
        return name2text, spk2name, name2spk

    def readWaveFiles(self, wav_files):
        name2wavpath = {}
        name2seg = {}
        
        dur_file = os.path.join(self.root_dir, "duration.json")
        if os.path.exists(dur_file):
            file2dur = json.load(open(dur_file))

        # # self.ReadMD5Thread(wav_files)
        for wav_file in tqdm(wav_files):
            filename = os.path.basename(wav_file)[:-4]
            filename = self.projname + "_" + filename
            name2wavpath[filename] = wav_file
            duration = file2dur[wav_file]
            name2seg[filename] = [0, duration]
        return name2wavpath, name2seg

    def process(self):
        txt_files = self.findOne(self.data_dir, "*.TXT")
        txt_files.extend(self.findOne(self.data_dir, "*.txt"))
        wav_files = self.findOne(self.data_dir, "*.WAV")
        wav_files.extend(self.findOne(self.data_dir, "*.wav"))
        print(f"Number of wave files: {len(txt_files)}...")
        print(f"Number of wave files: {len(wav_files)}...")
        print("=> Processing wave files...")
        name2wavpath, name2seg = self.readWaveFiles(wav_files)

        print("=> Processing text files...")
        name2text, spk2name, name2spk = self.readTextFile(txt_files, name2seg)
        print(f"=>> Number of name in text: {len(name2text.keys())}")

        self.splitAndWrite(spk2name, name2spk, name2text, name2wavpath)

    @property
    def info(self):
        info_dict = {}
        info_dict['raw_path'] = self.data_dir
        info_dict['kfolder'] = self.kformat_dir
        return info_dict

class InputBupaHKFormat(InputCollectBase):
    """
    For Customer CT-HK Long Audio Recordings.

    Example: demo.wav  demo.txt
        demo.wav: Long audio file
        demo.seg: Timestamps for the audio file with human speaking with transcribed text

    -----------------------------------
    Input Transcript: 
        @Customer:Agent
        [0.000] 喂
        [0.968] 喂 你好 唔該 劉 小姐 啊  
        [2.855] 我 係 啊  
        [3.580] 你好 我 係 kathy 陳 小姐 係 你好 係 以 下代 錄音 保障 返 你 而家 係咪 都 喺 香港
        [8.386] 係 佢 如果 做 嗱 而家 初步 你哋 問 嗰 啲 資料 啦 咁 到 最後 你哋 都 會  
        @Simultaneously
        [13.680] (?) 
    -----------------------------------
    Call the process method to get kaldi-format files for afterwards using.
    """
    def __init__(self, data_dir, kformat_dir, cut2short=False, mode="train", test_rate=0.001):
        super().__init__(data_dir, kformat_dir, cut2short, mode)
        self.name2files = self.find()
        print(f"Found {len(self.name2files)} Files...")

    def ReadTranscribedFile(self, path, duration):
        """
        Parser the transcribed file such as above example.
        """
        encoding_format = self.readEncodingType(path) # Get the encoding type. [utf-8] or [utf-8-sig] etc.
        cons = open(path, encoding=encoding_format, errors='ignore').readlines()

        text_parsers = []
        count = 0
        for line in cons:
            if line.startswith("@") or line.startswith("[End]"):
                continue
            line = line.strip()
            start_time = re.findall("\[([0-9]+\.[0-9]+)\]", line)[0]
            start_time = float(start_time)
            if len(text_parsers) != 0 and len(text_parsers[-1]) == 2:
                text_parsers[-1].append(start_time) # last seg's end time
            text = self.cleanText(line.split("]")[1])
            if self.check(text):
                print(text)
                continue
            if len(text) == 0: # no text remained after cleaning.
                continue
            text_parsers.append([text, start_time]) # current seg's start time
        if len(text_parsers) != 0 and len(text_parsers[-1]) == 2:
            text_parsers[-1].append(duration)
        return text_parsers


    def process(self):
        """
        Generate kaldi-format files.
        """
        names = []
        name2wavpath, name2spk  = {}, {}
        name2segments = defaultdict(list)

        for name, files in self.name2files.items():
            if len(files) != 2:
                continue
            wav_file, text_file = files
            self.ReadMD5(wav_file)
            try:
                duration = self.readDuration(wav_file)
            except Exception:
                print(wav_file)
                continue
            text_parsers = self.ReadTranscribedFile(text_file, duration) ## [ [ <text>, <start-time>, <end-time> ] ]
            if len(text_parsers) == 0:
                continue
            names.append(name)
            name2wavpath[name] = wav_file
            name2segments[name] = text_parsers
            name2spk[name] = name

        if self.mode == "train":
            test_nums = int(len(names) * self.test_rate)
            test_nums = max(test_nums, 1)
            train_nums = len(names) - test_nums
            
            train_name_file, test_name_file = os.path.join(self.kformat_dir, 'train.lst'), os.path.join(self.kformat_dir, 'test.lst')
            if os.path.exists(train_name_file) and os.path.exists(test_name_file) \
                and  os.path.getsize(train_name_file) and os.path.getsize(test_name_file):
                train_names = [ele.strip() for ele in open(train_name_file).readlines()]
                test_names = [ele.strip() for ele in open(test_name_file).readlines()]
            else:
                test_names = random.sample(names, test_nums)
                train_names = [name for name in names if name not in test_names]
                with open(train_name_file, 'w', encoding='utf-8') as ftr, \
                    open(test_name_file, 'w', encoding='utf-8') as fte:
                    for ele in train_names:
                        ftr.write(f"{ele}\n")
                    for ele in test_names:
                        fte.write(f"{ele}\n")

            train_kdir = os.path.join(self.kformat_dir, "train")
            test_kdir = os.path.join(self.kformat_dir, "test")

            os.makedirs(train_kdir, exist_ok=True)
            os.makedirs(test_kdir, exist_ok=True)

            self.write2Files(train_kdir, name2wavpath, name2segments, name2spk, train_names, long=True)
            self.write2Files(test_kdir, name2wavpath, name2segments, name2spk, test_names, long=True)
        else:
            os.makedirs(self.kformat_dir)
            self.write2Files(self.kformat_dir, name2wavpath, name2segments, name2spk, names, long=True)

    def find(self):
        """
        Find all .wav and .txt files and generate a dict with their path in a list
        Return:
            name2files: filename to [wav-path, trans-path]
        """
        name2files = defaultdict(list)
        wav_files = self.findOne(self.data_dir, "*.wav")
        txt_files = self.findOne(self.data_dir, "*.txt")
        
        for wav_file in wav_files:
            name = os.path.basename(wav_file)[:-4]
            name = name.replace(" ", "__")
            name2files[name].append(wav_file)

        for txt_file in txt_files:
            name = os.path.basename(txt_file)[:-4]
            name = name.replace(" ", "__")
            name2files[name].append(txt_file)

        return name2files

    @property
    def info(self):
        pass

class InputSOPhoneFormat(InputCollectBase):
    """
    For SpeechOcean Phone Recordings.

    Example: demo.txt.seg  demo.txt.wav demo.txt
        demo.wav: Long audio file
        demo.seg: Timestamps for the audio file with human speaking.

    Call the process method to get kaldi-format files for afterwards using.
    """
    def __init__(self, data_dir, kformat_dir, cut2short=True, lang="zh-CN"):
        super().__init__(data_dir, kformat_dir, cut2short)
        self.name2files = self.find()
        self.lang = lang
        print(f"Found {len(self.name2files.keys())} files...")

    def process(self):
        names = []
        spks = []
        name2wavpath, name2spk  = {}, {}
        name2text = {}
        os.makedirs(self.kformat_dir, exist_ok=True)
        for name, files in self.name2files.items():
            if len(files) != 3:
                continue
            wav_file, text_file, seg_file = files
            text_parsers = self.ReadTextAndSegFile(text_file, seg_file) ## [ [ <text>, *, <start-time>, <end-time> ] ]
            if len(text_parsers) == 0:
                continue
            if self.cut2short:
                subnames, subname2wavpath, subname2text, subname2spk = self.cut2ShortWave(wav_file, text_parsers, name)
                names.extend(subnames)
                name2wavpath.update(subname2wavpath)
                name2text.update(subname2text)
                name2spk.update(subname2spk)
            else:
                names.append(name)
                name2wavpath[name] = wav_file
                name2text[name] = text_parsers
                name2spk[name] = name

        if self.cut2ShortWave:
            long = False
        else:
            long = True

        self.splitAndWrite(spk2name, name2spk, name2text, name2wavpath)

    def cut2ShortWave(self, wav_file, text_parsers, name):
        sound = AudioSegment.from_wav(wav_file)
        subnames = []
        name2wavpath, name2text, name2spk = {}, {}, {}
        short_dir = os.path.join(os.path.dirname(self.kformat_dir), "short")
        os.makedirs(short_dir, exist_ok=True)
        for ele in text_parsers:
            text, segId, start_time, end_time = ele
            subname = f"{name}--{segId}--{start_time}--{end_time}"
            piece = sound[int(start_time*1000):int(end_time*1000)]
            subpath = os.path.join(short_dir, subname + ".wav")
            piece.export(subpath, format='wav')
            subnames.append(subname)
            name2wavpath[subname] = subpath
            name2text[subname] = text
            name2spk[subname] = name

        return subnames, name2wavpath, name2text, name2spk


    def ReadTextAndSegFile(self, txt_path, seg_path):

        encoding_format = self.readEncodingType(seg_path) # Get the encoding type. [utf-8] or [utf-8-sig] etc.
        seg_cons = open(seg_path, encoding=encoding_format, errors="ignore").readlines()[1:]
        text_parsers = []
        segId2timestamps = {}
        for line in seg_cons:
            line = line.strip()
            name, segId, *_, start_time, end_time, _, _ = line.split("<->")
            ## TODO Check if end_time is always larger then end_time. To filter the invalid data.
            start_time, end_time = int(start_time) / 1000, int(end_time) / 1000
            segId2timestamps[segId] = [start_time, end_time]

        encoding_format = self.readEncodingType(txt_path) 
        txt_cons = open(txt_path, encoding=encoding_format, errors="ignore").readlines()[1:]
        for line in txt_cons:
            line = line.strip()
            segId, text, *_ = line.split('<->')
            text = self.cleanText(text)
            if self.check(text):
                print(text)
                continue
            start_time, end_time = segId2timestamps[segId]
            text_parsers.append([text, segId, start_time, end_time])
        return text_parsers


    def check(self):
        """
        If there are not three files[.seg, .txt, .wav], it will be added to a dict.
        """
        error2Infos = defaultdict(list)
        for name, file_ in self.name2files.items():
            if len(file_) != 3:
                error2Infos["NumberError"].append([name, file_])
        return error2Infos


    def find(self):
        """
        Find all .wav and .seg files and .txt files and generate a dict with their path in a list
        Return:
            name2files: filename to [wav-path, txt-path, seg-path]
        """
        name2files = defaultdict(list)
        wav_files = self.findOne(self.data_dir, "*.wav")
        txt_files = self.findOne(self.data_dir, "*.txt")
        seg_files = self.findOne(self.data_dir, "*.seg")
        
        for wav_file in wav_files:
            name = os.path.basename(wav_file).split(".txt")[0]
            name2files[name].append(wav_file)

        for txt_file in txt_files:
            name = os.path.basename(txt_file).split(".txt")[0]
            name2files[name].append(txt_file)

        for seg_file in seg_files:
            name = os.path.basename(seg_file).split(".txt")[0]
            name2files[name].append(seg_file)

        return name2files

    @property
    def info(self):
        info_dict = {}
        info_dict['num_of_files'] = len(self.name2files.keys())
        info_dict['raw_path'] = self.data_dir
        info_dict['kfolder'] = self.kformat_dir
        return info_dict

class InputKingASR118(InputCollectBase):
    def __init__(self, data_dir, kformat_dir, mode="train", test_rate=0.001, device="Android", lang="zh-CN"):
        self.root_dir = data_dir
        self.projname = os.path.basename(data_dir)
        self.data_dir = os.path.join(data_dir, device)
        self.kformat_dir = kformat_dir
        self.mode = mode
        self.test_rate = test_rate
        self.lang = lang

    def readTextFile(self, txt_files, name2seg):
        name2text = {}
        spk2name = defaultdict(list)
        name2spk = {}
        for txt_file in txt_files:
            one, two = txt_file.split("/")[-2:]
            filename = one + "--" + two[:-4]
            filename = self.projname + "_" + filename
            encoding_format = self.readEncodingType(txt_file)
            contents = open(txt_file, encoding=encoding_format).readlines()
            for index in range(0, len(contents), 3):
                name = contents[index].strip().split("\t")[0]
                name = filename + "--" + name
                try:
                    start_time, end_time = name2seg[name]
                except KeyError:
                    continue
                spk2name[filename].append(name)
                name2spk[name] = filename
                text = contents[index+2].strip().split("\t")[-1]
                text = self.cleanText(text)
                if self.check(text):
                    print(text)
                    continue
                name2text[name] = [[text, start_time, end_time]]
        return name2text, spk2name, name2spk

    def readWaveFiles(self, wav_files):
        name2wavpath = {}
        name2seg = {}
        # self.ReadMD5Thread(wav_files)
        for wav_file in tqdm(wav_files):
            one, two, three = wav_file.split("/")[-3:]
            filename = one+"--" + two + "--" + three[:-4]
            filename = self.projname + "_" + filename
            name2wavpath[filename] = wav_file
            duration = self.readDuration(wav_file)
            name2seg[filename] = [0, duration]
        return name2wavpath, name2seg

    def process(self):
        txt_files = self.findOne(self.data_dir, "*.txt")
        wav_files = self.findOne(self.data_dir, "*.wav")

        print("=> Processing wave files...")
        name2wavpath, name2seg = self.readWaveFiles(wav_files)

        print("=> Processing text files...")
        name2text, spk2name, name2spk = self.readTextFile(txt_files, name2seg)

        self.splitAndWrite(spk2name, name2spk, name2text, name2wavpath)


    @property
    def info(self):
        pass

class InputKingASR214(InputKingASR118):
    def __init__(self, data_dir, kformat_dir, mode="train", test_rate=0.001, device="Android", lang="zh-CN"):
        self.projname = os.path.basename(data_dir)
        self.data_dir = os.path.join(data_dir, "DATA")
        self.kformat_dir = kformat_dir
        self.mode = mode
        self.device = device
        self.test_rate = test_rate
        self.lang = lang

    def readWaveFiles(self, wav_files):
        name2wavpath = {}
        name2seg = {}
        # self.ReadMD5Thread(wav_files)
        for wav_file in tqdm(wav_files):
            one, two, _, three = wav_file.split("/")[-4:]
            two = f"Speaker{two}"
            filename = "script--" + two + "--" + three[:-4]
            filename = self.projname + "_" + filename
            name2wavpath[filename] = wav_file
            duration = self.readDuration(wav_file)
            name2seg[filename] = [0, duration]
        return name2wavpath, name2seg

    def process(self):
        wav_dir = os.path.join(self.data_dir, "wave", self.device)
        txt_dir = os.path.join(self.data_dir, "script")

        txt_files = self.findOne(txt_dir, "*.txt")
        wav_files = self.findOne(wav_dir, "*.wav")

        print("=> Processing wave files...")
        name2wavpath, name2seg = self.readWaveFiles(wav_files)

        print("=> Processing text files...")
        name2text, spk2name, name2spk = self.readTextFile(txt_files, name2seg)

        self.splitAndWrite(spk2name, name2spk, name2text, name2wavpath)

    @property
    def info(self):
        pass

class InputTR2020063(InputCollectBase):
    """
    ------
    A:male
    B:female
    C:male
    ------
    [00:00:00,00:00:01.814,A]人会怎么看这个新发型？
    [00:00:01.998,00:00:03.698,B]他们会大吃一 /n:laughter 惊。
    """
    def __init__(self, data_dir, kformat_dir, mode="train", test_rate=0.001, lang="zh-CN"):
        self.data_dir = data_dir
        self.kformat_dir = kformat_dir
        self.mode = mode
        self.test_rate = test_rate
        self.projname = os.path.basename(data_dir)
        self.lang = lang
        self.info_dict = {}

    def timeTranslator(self, ftime):
        try:
            h, m, s = ftime.split(":")
        except ValueError:
            m, s = ftime.split(":")
            h = 0
        try:
            h, m, s = float(h), float(m), float(s)
        except ValueError:
            return -1
        duration = h * 3600 + m * 60 + s 
        return duration

    def readTextFile(self, text_files):
        name2text = defaultdict(list)
        spk2name = defaultdict(list)
        name2spk = {}
        duration_total = 0
        textnames = defaultdict(list)
        for text_file in tqdm(text_files):
            filename = os.path.basename(text_file)[:-4]
            if self.projname == "TR2021041":
                filename = filename.split("-", 2)[-1]
            filename = self.projname + "_" + filename
            textnames[filename].append(text_file)
            encoding_format = self.readEncodingType(text_file)
            cons = open(text_file, encoding=encoding_format, errors="ignore").readlines()
            for line in cons:
                line = line.strip()
                if not line.startswith("["):
                    continue
                text = line.split("]")[-1]
                text = re.sub("/.*?/", "", text)
                text = re.sub("/[a-zA-Z]+:?(?:[a-zA-Z]+)?-?(?:[a-zA-Z]+)?(?:[0-9]+)?", " ", text)
                if self.projname in ["TR2020063", "TR2020186", "TR2020107"]:
                    if not text.startswith("spk "):
                        continue
                text = text.replace("spk ", "").replace("·", "")
                text = self.cleanText(text)
                text = text.replace("spk_noise", "").replace("spk_bg", "").replace("spk_unknown", "")
                if not text.replace(" ", ""):
                    continue
                if self.check(text):
                    continue
                # [00:00:04.527,00:00:06.383,A]
                start_time, end_time = re.findall("\[(.*?)\]", line)[0].split(",")[:2]
                start_time, end_time = self.timeTranslator(start_time), self.timeTranslator(end_time)
                if end_time <= start_time:
                    print(text_file, line)
                    continue
                name2text[filename].append([text, start_time, end_time])
                duration_total += end_time - start_time
            spk2name[filename].append(filename)
            name2spk[filename] = filename
        print("Total Duration: ", duration_total)
        return name2text, spk2name, name2spk

    def process(self):
        trtext_dir = os.path.join(self.data_dir, "product")
        trwave_dir = os.path.join(self.data_dir, "wave_org")
        text_files = self.findOne(trtext_dir, "*.txt")
        name2text, spk2name, name2spk = self.readTextFile(text_files)

        wav_files = self.findOne(trwave_dir, "*.wav")
        
        self.info_dict["num_textgrid_files"] = len(text_files)
        self.info_dict["num_wave_files"] = len(wav_files)
        
        name2wavpath = {}
        # self.ReadMD5Thread(wav_files)
        for wav_file in tqdm(wav_files):
            filename = os.path.basename(wav_file)[:-4]
            if self.projname in ["TR2020114"]:
                filename = filename.split("-")[-1]
            if self.projname in ["TR2020146"]:
                filename = filename.replace(":", "_")
            filename = self.projname + "_" + filename
            name2wavpath[filename] = wav_file
        self.splitAndWrite(spk2name, name2spk, name2text, name2wavpath)

class InputVIPKIDTR2020064(InputCollectBase):
    """
    excel Text
    双通道，先将语音转换成单通道，后接_0和_1
    """
    def __init__(self, data_dir, kformat_dir, mode="train", test_rate=0.001, lang="zh-CN"):
        self.data_dir = data_dir
        self.kformat_dir = kformat_dir
        self.mode = mode
        self.test_rate = test_rate
        self.projname = os.path.basename(data_dir)
        self.lang = lang
        self.info_dict = {}

    def readxlsxFile(self, xlsx_files):
        name2text = defaultdict(list)
        spk2name = defaultdict(list)
        name2spk = {}
        duration_total = 0
        for xlsx_file in tqdm(xlsx_files):
            filename = os.path.basename(xlsx_file)[:-5]
            wb_obj = openpyxl.load_workbook(xlsx_file)
            sheet_obj = wb_obj.active
            for row in range(2, sheet_obj.max_row + 1):
                start_time = sheet_obj.cell(row = row, column = 2).value
                end_time = sheet_obj.cell(row = row, column = 3).value
                cid = str(sheet_obj.cell(row = row, column = 4).value)
                text = sheet_obj.cell(row = row, column = 6).value
                # [NUM#四#4]
                text = re.sub("\[.*?#(.*?)#.*?\]", lambda x: x.groups()[0], text).replace("-", " ")
                text = self.cleanText(text)
                if not text.replace(" ", ""):
                    continue
                if self.check(text):
                    print(text)
                    continue
                cfilename = self.projname + "_" + filename + "_" + cid
                name2text[cfilename].append([text, int(start_time)/1000, int(end_time)/1000])
                duration_total += int(end_time)/1000 - int(start_time)/1000
                spk2name[cfilename] = [cfilename]
                name2spk[cfilename] = cfilename


        print("Total Duration: ", duration_total)
        return name2text, spk2name, name2spk


    def process(self):
        trtext_dir = os.path.join(self.data_dir, "product")
        trwave_dir = os.path.join(self.data_dir, "wave_org")

        text_files = self.findOne(trtext_dir, "*.xlsx")
        name2text, spk2name, name2spk = self.readxlsxFile(text_files)

        wav_files = self.findOne(trwave_dir, "*.wav")
        
        self.info_dict["num_xlsx_files"] = len(text_files)
        self.info_dict["num_wave_files"] = len(wav_files)

        name2wavpath = {}
        # self.ReadMD5Thread(wav_files)
        for wav_file in wav_files:
            filename = self.projname + "_" + os.path.basename(wav_file)[:-4]
            name2wavpath[filename] = wav_file

        self.splitAndWrite(spk2name, name2spk, name2text, name2wavpath)

class InputStrangeTGTR2019006(InputCollectBase):
    """
    奇奇怪怪的TextGrid格式  

    """
    def __init__(self, data_dir, kformat_dir, mode="train", test_rate=0.001, lang="zh-CN"):
        self.data_dir = data_dir
        self.kformat_dir = kformat_dir
        self.mode = mode
        self.test_rate = test_rate
        self.projname = os.path.basename(data_dir)
        self.lang = lang
        self.info_dict = {}

    def readTextGridFile(self, text_files):
        name2text = defaultdict(list)
        spk2name = defaultdict(list)
        name2spk = {}
        duration_total = 0
        
        for text_file in text_files:
            filename = os.path.basename(text_file)[:-9]
            filename = self.projname + "_" + filename
            encoding_format = self.readEncodingType(text_file)
            cons = open(text_file, encoding=encoding_format, errors="ignore").read().split('"CONTENT"\n')[-1].split("\n")[3:]
            for index in range(0, len(cons), 3):
                try:
                    start_time, end_time, text = cons[index:index+3]
                    start_time, end_time = float(start_time), float(end_time)
                except ValueError:
                    continue
                if not re.findall(r"[\u4e00-\u9fa5]", text):
                    continue
                text = text.replace('"', '').replace("+", "")
                text = self.cleanText(text)
                if not text.replace(" ", ""):
                    continue
                if self.check(text):
                    print(text)
                    continue
                name2text[filename].append([text, start_time, end_time])
                duration_total += end_time - start_time
            name2spk[filename] = filename
            spk2name[filename].append(filename)
        print("Total Duration[h]: ", duration_total / 3600)
        return name2text, spk2name, name2spk


    def process(self):
        trtext_dir = os.path.join(self.data_dir, "product")
        trwave_dir = os.path.join(self.data_dir, "wave_org")

        text_files = self.findOne(trtext_dir, "*.TextGrid")
        name2text, spk2name, name2spk = self.readTextGridFile(text_files)

        wav_files = self.findOne(trwave_dir, "*.wav")
        
        self.info_dict["num_text_files"] = len(text_files)
        self.info_dict["num_wave_files"] = len(wav_files)

        name2wavpath = {}
        # self.ReadMD5Thread(wav_files)
        for wav_file in wav_files:
            filename = self.projname + "_" + os.path.basename(wav_file)[:-4]
            name2wavpath[filename] = wav_file

        self.splitAndWrite(spk2name, name2spk, name2text, name2wavpath)

class InputTRTextGrid(InputCollectBase):
    """
    转写的Textgrid格式 [TR2019068]
    """
    def __init__(self, data_dir, kformat_dir, mode="train", test_rate=0.001, item_index=0, lang="zh-CN"):
        self.data_dir = data_dir
        self.kformat_dir = kformat_dir
        self.mode = mode
        self.test_rate = test_rate
        self.projname = os.path.basename(data_dir)
        self.info_dict = {}
        self.item_index = item_index
        self.lang = lang

    def readTextGridFile(self, tg_files):
        name2text = defaultdict(list)
        spk2name = defaultdict(list)
        name2spk = {}
        duration_total = 0
        for tg_file in tqdm(tg_files):
            if "yueyu" in tg_file:
                continue
            if self.projname == "TR2019068":
                filename =  os.path.basename(os.path.dirname(tg_file)).replace("result-", "") + "--" + os.path.basename(tg_file)[:-9]
            else:
                filename = os.path.basename(tg_file)[:-9]
            filename = self.projname + "_" + filename
            tg = TextGrid()
            tg.read(tg_file)
            try:
                for info in tg.tiers:
                    if info.name == "PEOPLE" or info.name == "text":
                        break
                if info.name != "PEOPLE" and info.name != "text":
                    info = tg.tiers[self.item_index]
                # if info.name == "SPEAKER":
                #     info = tg.tiers[1] ## BUG
            except Exception:
                info = tg.tiers[0]
            for ele in info:
                text, maxt, mint = ele.mark, float(ele.maxTime), float(ele.minTime)
                if (text == "sil" or text == "noi") and self.projname == "TR2019068":
                    continue
                text = self.cleanText(text)
                if self.check(text):
                    print(text)
                    continue
                if len(text) == 0:
                    continue
                duration_total += maxt - mint
                name2text[filename].append([text, mint, maxt])
            name2spk[filename] = filename
            spk2name[filename].append(filename)
        print("Total Duration[h]: ", duration_total / 3600)
        return name2text, spk2name, name2spk


    def process(self):
        trtext_dir = os.path.join(self.data_dir, "product")
        trwave_dir = os.path.join(self.data_dir, "wave_org")

        text_files = self.findOne(trtext_dir, "*.TextGrid")
        name2text, spk2name, name2spk = self.readTextGridFile(text_files)

        wav_files = self.findOne(trwave_dir, "*.wav")
        
        self.info_dict["num_text_files"] = len(text_files)
        self.info_dict["num_wave_files"] = len(wav_files)

        name2wavpath = {}
        name2wavnums = defaultdict(lambda : 0)
        # self.ReadMD5Thread(wav_files)
        for wav_file in tqdm(wav_files):
            if self.projname == "TR2019068":
                filename = os.path.basename(os.path.dirname(wav_file)) + "--" + os.path.basename(wav_file)[:-4]
            else:
                filename = os.path.basename(wav_file)[:-4]
            filename = self.projname + "_" + filename
            name2wavnums[filename] += 1
            name2wavpath[filename] = wav_file

        for name, nums in name2wavnums.items():
            if nums > 1:
                print(name)
                print(name2wavpath[name])
                # sys.exit()

        self.splitAndWrite(spk2name, name2spk, name2text, name2wavpath)

class InputResultTxtType0(InputCollectBase):
    """
    Example: demo.wav result.txt invalid_result.txt [TR2019131、TR2019132、TR2020018]
            音频名  文本 
            9ff3ef4a9c79479c931a1bbfc6c2072b    钥匙的钥可以组什么词 
            e4fa290f350e4c1a97a6430be96d8739    在银行上一二三四五的大写怎么写 
            7a59e5f2ea024cc0b7d86948518f6ddb    CAN SPEAK ENGLISH
            20f20d3afb0a4b14a02703b3921ebe28    狗狗积食了不吃东西有利热怎么办 
            6e7ceb4271b74a1395a899f3d5b360ed    青岛的网红店有哪些 
    或者: 
            声音    文本    是否有效        底噪    口音
            587dd25c25c6577508dd184f12501.WAV       球星玩偶        有效    有      无
            5890b484fd079fcea5eff3b639552.WAV       带灯的挖耳勺    有效    有      无
            5890b495c38aeb983eb07bc828995.WAV       怡宝    有效    有      无
    """
    def __init__(self, data_dir, kformat_dir, mode="train", test_rate=0.001, lang="zh-CN"):
        self.data_dir = data_dir
        self.projname = os.path.basename(data_dir)
        self.kformat_dir = kformat_dir
        self.mode = mode
        self.test_rate = test_rate
        self.lang = lang

    def cleanText(self, text):
        """
        Clean text to remove invailable tags or punctuations if needed!
        """
        text = self.strQ2B(text)        
        text = re.sub("【.*?】", "", text)
        if self.projname not in ["TR2019152"]:
            text = re.sub("\[.*?\]", '', text)
        else:
            text = text.replace("[", "").replace("]", "")
        text = re.sub('<.*?>', ' ', text)
        text = re.sub('\(.*?\)', ' ', text)
        text = re.sub('[，。？,；、`！──－：‘’“”~《》．"；,/×—－─（）・／—"『〈〉』·「」<>.?!…;]', ' ', text)
        text = re.sub("--'([\u4e00-\u9fa5]+)'", lambda x:x.groups()[0], text)
        text = text.replace('*', ' ')
        text = " ".join(text.split())
        if "{" in text: # 有误读的 大括号里面写的还是正常的句子，这种读错的我们不要
            text = ""   # 例如 我 看 真正 的 安定团结 是 肯定 的 ， 现在 是 说 日 <B>{渐gai4}</B> 巩固
        return text


    def readTextFile(self, txt_files, name2wavpath):
        self.name2text = {}
        self.spk2name = defaultdict(list)
        self.name2spk = {}
        for txt_file in txt_files:
            if "invalid" in txt_file or "attrs" in txt_file:
                continue
            print(f"=> {txt_file}...")
            encoding_format = self.readEncodingType(txt_file)
            contents = open(txt_file, encoding=encoding_format).readlines()[1:]
            nums_perpart = len(contents) // 30 + 1
            parts = [contents[l:l+nums_perpart] for l in range(0, len(contents), nums_perpart)]
            print(len(contents))
            threads = []
            for part in parts:
                t = threading.Thread(target=self.readTextPer, args=(part, name2wavpath,))
                threads.append(t)
                t.start()
            for t in threads:
                t.join()

        return self.name2text, self.spk2name, self.name2spk

    def readTextPer(self, contents, name2wavpath):
        for index, line in enumerate(contents):
            if index % 1000 == 0:
                print(f"[ {index} / {len(contents)} ] Reading Text Files...")
            try:
                filename, text = line.strip().split("\t")[:2]
            except Exception:
                print(line)
                continue
            if len(filename) < 10 or self.projname in ["TR2017067", "TR20170023"]:
                filename = "___".join(filename.split("\\")[-3:])
            else:
                filename = filename.split("\\")[-1]
            if self.projname in ["TR2017038", "TR2020192", "TR2019025", "TR2020018", "TR2020037", "TR2017060", "TR20170023", "TR2017100", "TR2017072", "TR2017086", "TR2020018", "TR2017067", "TR2019038", "TR2018134", "TR2017085", "TR2017120", "TR2018117", "TR2018134", "TR2017108", "RS-TR-68", "TS2019032", "TR2018096"]:
                filename = filename[:-4]
            filename = self.projname + "_" + filename
            try:
                filepath = name2wavpath[filename]
            except KeyError:
                print("WAV NOT FOUND: ", filename)
                continue
            try:
                duration = self.readDuration(filepath)
            except Exception:
                print("AUDIO READ FAILED!")
                print(filename)
                continue
            text = re.sub("/[a-zA-Z]+:?-?(?:[a-zA-Z]+)?([0-9]+)?", " ", text)
            text = self.cleanText(text)
            text = text.replace("@", " at ")
            if self.check(text):
                print(text)
                continue
                # sys.exit()
            self.name2text[filename] = [[text, 0, duration]]
            self.spk2name[filename].append(filename)
            self.name2spk[filename] = filename


    def readWaveFiles(self, wav_files):
        name2wavpath = {}

        self.writeDurationThread(wav_files)
        # self.ReadMD5Thread(wav_files)
        for wav_file in tqdm(wav_files):
            filename = os.path.basename(wav_file)[:-4]
            if self.projname in ["TR20170023", "TR2017067"]:
                one = os.path.dirname(wav_file)
                two = os.path.basename(one)
                three = os.path.basename(os.path.dirname(one))
                filename = "___".join([three, two, filename])
            filename = self.projname + "_" + filename
            name2wavpath[filename] = wav_file
        return name2wavpath

    def process(self):
        trtext_dir = os.path.join(self.data_dir, "product")
        trwave_dir = os.path.join(self.data_dir, "wave_org")
        print(trtext_dir)
        print(trwave_dir)
        txt_files = self.findOne(trtext_dir, "*result*.txt")
        print(f"Found text file: {len(txt_files)}...")
        wav_files = self.findOne(trwave_dir, "*.wav")
        wav_files.extend(self.findOne(trwave_dir, "*.WAV"))
        print(f"Found wave file: {len(wav_files)}...")
        print("=> Processing wave files...")
        name2wavpath = self.readWaveFiles(wav_files)
        print("=> Processing text files...")
        name2text, spk2name, name2spk = self.readTextFile(txt_files, name2wavpath)

        self.splitAndWrite(spk2name, name2spk, name2text, name2wavpath)

    @property
    def info(self):
        pass

class InputResultTxtType1(InputCollectBase):
    """
    Example: demo.wav result.txt invalid_result.txt 
    编号    语音文件名      文本    性别    有无口音        是否清晰
    19827   151012323_9584470C.ogg  第一个  男      无      是
    """
    def __init__(self, data_dir, kformat_dir, mode="train", test_rate=0.001, lang="zh-CN"):
        self.data_dir = data_dir
        self.projname = os.path.basename(data_dir)
        self.kformat_dir = kformat_dir
        self.mode = mode
        self.test_rate = test_rate
        self.lang = lang

    def readTextFile(self, txt_files, name2wavpath):
        self.name2text = {}
        self.spk2name = defaultdict(list)
        self.name2spk = {}
        for txt_file in txt_files:
            if "invalid" in txt_file:
                continue
            print(f"=> {txt_file}...")
            encoding_format = self.readEncodingType(txt_file)
            contents = open(txt_file, encoding=encoding_format).readlines()[1:]
            nums_perpart = len(contents) // 30 + 1
            parts = [contents[l:l+nums_perpart] for l in range(0, len(contents), nums_perpart)]
            print(len(contents))
            threads = []
            for part in parts:
                t = threading.Thread(target=self.readTextPer, args=(part, name2wavpath,))
                threads.append(t)
                t.start()
            for t in threads:
                t.join()
        return self.name2text, self.spk2name, self.name2spk
                
    def readTextPer(self, contents, name2wavpath):
        for index, line in enumerate(contents):
            if index % 1000 == 0:
                print(f"[ {index} / {len(contents)} ] Reading Text Files...")
            filename, text = line.strip().split("\t")[1:3]
            filename = filename[:-4]
            filename = self.projname + "_" + filename
            try:
                filepath = name2wavpath[filename]
            except KeyError:
                print(filename)
                continue
            try:
                duration = self.readDuration(filepath)
            except Exception as e:
                print(e)
                print(filename)
                # sys.exit()
                continue
            text = re.sub("/[a-zA-Z]+:?-?(?:[a-zA-Z]+)?([0-9]+)?", " ", text)
            text = self.cleanText(text)
            if self.check(text):
                print(text)
                continue
                # sys.exit()
            self.name2text[filename] = [[text, 0, duration]]
            self.spk2name[filename].append(filename)
            self.name2spk[filename] = filename

    def readWaveFiles(self, wav_files):
        name2wavpath = {}

        # self.writeDurationThread(wav_files)
        # self.ReadMD5Thread(wav_files)
        for wav_file in tqdm(wav_files):
            filename = os.path.basename(wav_file)[:-4]
            filename = self.projname + "_" + filename
            name2wavpath[filename] = wav_file
        return name2wavpath

    def process(self):
        trtext_dir = os.path.join(self.data_dir, "product")
        trwave_dir = os.path.join(self.data_dir, "wave_org")
        print(trtext_dir)
        print(trwave_dir)
        txt_files = self.findOne(trtext_dir, "*result.txt")
        print(txt_files)
        print(f"Found text file: {len(txt_files)}...")
        wav_files = self.findOne(trwave_dir, "*.wav")
        print(f"Found wave file: {len(wav_files)}...")
        
        print("=> Processing wave files...")
        name2wavpath = self.readWaveFiles(wav_files)

        print("=> Processing text files...")
        name2text, spk2name, name2spk = self.readTextFile(txt_files, name2wavpath)
        self.splitAndWrite(spk2name, name2spk, name2text, name2wavpath)

    @property
    def info(self):
        pass

class InputTransFormatLBLShort(InputCollectBase):
    """
    Example: demo.wav result.txt invalid_result.txt [TR2019131、TR2019132、TR2020018]
            音频名  文本 
            9ff3ef4a9c79479c931a1bbfc6c2072b    钥匙的钥可以组什么词 
            e4fa290f350e4c1a97a6430be96d8739    在银行上一二三四五的大写怎么写 
            7a59e5f2ea024cc0b7d86948518f6ddb    CAN SPEAK ENGLISH
            20f20d3afb0a4b14a02703b3921ebe28    狗狗积食了不吃东西有利热怎么办 
            6e7ceb4271b74a1395a899f3d5b360ed    青岛的网红店有哪些 
    或者: 
            声音    文本    是否有效        底噪    口音
            587dd25c25c6577508dd184f12501.WAV       球星玩偶        有效    有      无
            5890b484fd079fcea5eff3b639552.WAV       带灯的挖耳勺    有效    有      无
            5890b495c38aeb983eb07bc828995.WAV       怡宝    有效    有      无
    """
    def __init__(self, data_dir, kformat_dir, mode="train", test_rate=0.001, lang="zh-CN"):
        self.data_dir = data_dir
        self.projname = os.path.basename(data_dir)
        self.kformat_dir = kformat_dir
        self.mode = mode
        self.test_rate = test_rate

    def cleanText(self, text):
        """
        Clean text to remove invailable tags or punctuations if needed!
        """
        text = re.sub("【.*?】", "", text)
        if self.projname not in ["TR2019152"]:
            text = re.sub("\[.*?\]", '', text)
        else:
            text = text.replace("[", "").replace("]", "")
        text = re.sub('<.*?>', ' ', text)
        text = re.sub('\(.*?\)', ' ', text)
        text = re.sub('[，。？,；、`！──－：‘’“”~《》．"；,/×—－─（）・／—"『〈〉』·「」<>.?!…;]', '', text)
        text = re.sub("--'([\u4e00-\u9fa5]+)'", lambda x:x.groups()[0], text)
        text = text.replace('*', ' ')
        text = " ".join(text.split())
        if "{" in text: # 有误读的 大括号里面写的还是正常的句子，这种读错的我们不要
            text = ""   # 例如 我 看 真正 的 安定团结 是 肯定 的 ， 现在 是 说 日 <B>{渐gai4}</B> 巩固
        text = self.strQ2B(text)
        return text


    def readTextFile(self, txt_files, name2wavpath):
        name2text = {}
        spk2name = defaultdict(list)
        name2spk = {}
        for txt_file in tqdm(txt_files):
            try:
                encoding_format = self.readEncodingType(txt_file)
            except Exception:
                continue
            line = open(txt_file, encoding=encoding_format).readlines()[-2]
            if "annotated text" not in line:
                continue
            line = line.strip()
            filename = os.path.basename(txt_file)[:-4]
            text = line.split(":")[-1]
            
            filename = self.projname + "_" + filename
            try:
                filepath = name2wavpath[filename]
            except KeyError:
                print("WAV NOT FOUND: ", filename)
                continue
            try:
                duration = self.readDuration(filepath)
            except Exception:
                print("AUDIO READ FAILED!")
                print(filename)
                continue
            text = self.cleanText(text)
            text = text.replace("@", " at ")
            if self.check(text):
                print(text)
                continue
                # sys.exit()
            name2text[filename] = [[text, 0, duration]]
            spk2name[filename].append(filename)
            name2spk[filename] = filename
        return name2text, spk2name, name2spk

    def readWaveFiles(self, wav_files):
        name2wavpath = {}

        self.writeDurationThread(wav_files)
        # self.ReadMD5Thread(wav_files)
        for wav_file in tqdm(wav_files):
            filename = os.path.basename(wav_file)[:-4]
            if self.projname in ["TR20170023", "TR2017067"]:
                one = os.path.dirname(wav_file)
                two = os.path.basename(one)
                three = os.path.basename(os.path.dirname(one))
                filename = "___".join([three, two, filename])
            filename = self.projname + "_" + filename
            name2wavpath[filename] = wav_file
        return name2wavpath

    def process(self):
        trtext_dir = os.path.join(self.data_dir, "product")
        trwave_dir = os.path.join(self.data_dir, "wave_org")
        print(trtext_dir)
        print(trwave_dir)
        txt_files = self.findOne(trtext_dir, "*.lbl")
        print(f"Found text file: {len(txt_files)}...")
        wav_files = self.findOne(trwave_dir, "*.wav")
        # wav_files.extend(self.findOne(trwave_dir, "*.WAV"))
        print(f"Found wave file: {len(wav_files)}...")
        print("=> Processing wave files...")
        name2wavpath = self.readWaveFiles(wav_files)
        print("=> Processing text files...")
        name2text, spk2name, name2spk = self.readTextFile(txt_files, name2wavpath)

        self.splitAndWrite(spk2name, name2spk, name2text, name2wavpath)

    @property
    def info(self):
        pass

class InputTransFormatXLSXShort(InputCollectBase):
    """
    ファイル名	書き起こし/#NG	有効/無効タグ	方言
    wave/000480-0000.wav	#NG	NN	
    wave/000480-0001.wav	欢迎光临，你好。	OK	
    wave/000480-0002.wav	你在干嘛？	OK	
    wave/000480-0003.wav	我们只差十元就可以出去了是吗？	OK	

    """
    def __init__(self, data_dir, kformat_dir, mode="train", test_rate=0.001, lang="zh-CN"):
        self.data_dir = data_dir
        self.projname = os.path.basename(data_dir)
        self.kformat_dir = kformat_dir
        self.mode = mode
        self.test_rate = test_rate
        self.lang = lang

    def readTextFile(self, xlsx_files, name2wavpath):
        name2text = {}
        spk2name = defaultdict(list)
        name2spk = {}
        for xlsx_file in tqdm(xlsx_files):
            filename = os.path.basename(xlsx_file)[:-5]
            wb_obj = openpyxl.load_workbook(xlsx_file)
            sheet_obj = wb_obj.active
            for row in range(2, sheet_obj.max_row + 1):
                filename = sheet_obj.cell(row = row, column = 1).value
                text = str(sheet_obj.cell(row = row, column = 2).value)
                vflag = sheet_obj.cell(row = row, column = 3).value
                if vflag != "OK":
                    continue
                try:
                    filename = re.findall('\("(.*?)"\)', filename)[0]
                    filename = self.projname + "_" + os.path.basename(filename)[:-4]
                except IndexError:
                    print(filename)
                    sys.exit()
                text = self.cleanText(text)
                if not text.replace(" ", ""):
                    continue
                if self.check(text):
                    print(text)
                    continue
                try:
                    filepath = name2wavpath[filename]
                except Exception:
                    print(filename)
                    continue
                try:
                    duration = self.readDuration(filepath)
                except Exception:
                    print("AUDIO READ FAILED!")
                    print(filename)
                    continue
                name2text[filename] = [[text, 0, duration]]
                spk2name[filename].append(filename)
                name2spk[filename] = filename
        return name2text, spk2name, name2spk

    def readWaveFiles(self, wav_files):
        name2wavpath = {}
        self.writeDurationThread(wav_files)
        # self.ReadMD5Thread(wav_files)
        for wav_file in tqdm(wav_files):
            filename = os.path.basename(wav_file)[:-4]
            filename = self.projname + "_" + filename
            name2wavpath[filename] = wav_file
        return name2wavpath

    def process(self):
        trtext_dir = os.path.join(self.data_dir, "product")
        trwave_dir = os.path.join(self.data_dir, "wave_org")
        print(trtext_dir)
        print(trwave_dir)
        txt_files = self.findOne(trtext_dir, "*.xlsx")
        print(f"Found text file: {len(txt_files)}...")
        wav_files = self.findOne(trwave_dir, "*.wav")
        wav_files.extend(self.findOne(trwave_dir, "*.WAV"))

        print(f"Found wave file: {len(wav_files)}...")
        print("=> Processing wave files...")
        name2wavpath = self.readWaveFiles(wav_files)
        print("=> Processing text files...")
        name2text, spk2name, name2spk = self.readTextFile(txt_files, name2wavpath)

        self.splitAndWrite(spk2name, name2spk, name2text, name2wavpath)

    @property
    def info(self):
        pass

class InputTransFormatXLSXLong(InputCollectBase):
    """
    编号	文本	是否有效	时段
    612352_1	到达直播间同学来我们发一个准点红包好吗？	有效	[19.041][22.366]
    612352_2	好领一下我们今天的准点红包，好的。	有效	[25.525][29.237]
    612352_3	好，来看看今天的幸运星和倒霉蛋哈。	有效	[30.448][34.523]
    612352_4	梓乐幸运星二十九枚，还有我们的帅帅二十九枚。	有效	[34.523][38.325]
    612352_5	昊阳二十九枚，啊以及若夕二十九枚不错。	有效	[38.325][42.223]
    """
    def __init__(self, data_dir, kformat_dir, mode="train", test_rate=0.001, lang="zh-CN"):
        self.data_dir = data_dir
        self.projname = os.path.basename(data_dir)
        self.kformat_dir = kformat_dir
        self.mode = mode
        self.test_rate = test_rate
        self.lang = lang

    def readTextFile(self, xlsx_files):
        name2text = defaultdict(list)
        spk2name = defaultdict(list)
        name2spk = {}
        total_dur = 0
        for xlsx_file in tqdm(xlsx_files):
            filename = os.path.basename(xlsx_file)[:-5]
            filename = self.projname + "_" + filename
            wb_obj = openpyxl.load_workbook(xlsx_file)
            sheet_obj = wb_obj.active
            for row in range(2, sheet_obj.max_row + 1):
                text = str(sheet_obj.cell(row = row, column = 2).value)
                vflag = sheet_obj.cell(row = row, column = 3).value
                timeinfo = str(sheet_obj.cell(row = row, column = 4).value)
                if vflag != "有效":
                    continue
                try:
                    start_time, end_time = re.findall("\[(.*?)\]", timeinfo)
                    start_time, end_time = float(start_time), float(end_time)
                    if end_time <= start_time:
                        print(text_file, timeinfo)
                        continue
                except Exception:
                    print(timeinfo)
                    continue
                total_dur += end_time - start_time
                text = self.cleanText(text)
                if not text.replace(" ", ""):
                    continue
                if self.check(text):
                    print(text)
                    continue
                name2text[filename].append([text, start_time, end_time])
            spk2name[filename].append(filename)
            name2spk[filename] = filename
        print("Total Duration: ", total_dur/3600)
        return name2text, spk2name, name2spk

    def readWaveFiles(self, wav_files):
        name2wavpath = {}
        # self.ReadMD5Thread(wav_files)
        for wav_file in tqdm(wav_files):
            filename = os.path.basename(wav_file)[:-4]
            filename = self.projname + "_" + filename
            name2wavpath[filename] = wav_file
        return name2wavpath

    def process(self):
        trtext_dir = os.path.join(self.data_dir, "product")
        trwave_dir = os.path.join(self.data_dir, "wave_org")
        print(trtext_dir)
        print(trwave_dir)
        txt_files = self.findOne(trtext_dir, "*.xlsx")
        print(f"Found text file: {len(txt_files)}...")
        wav_files = self.findOne(trwave_dir, "*.wav")
        wav_files.extend(self.findOne(trwave_dir, "*.WAV"))

        print(f"Found wave file: {len(wav_files)}...")
        print("=> Processing wave files...")
        name2wavpath = self.readWaveFiles(wav_files)
        print(name2wavpath)
        print("=> Processing text files...")
        name2text, spk2name, name2spk = self.readTextFile(txt_files)

        self.splitAndWrite(spk2name, name2spk, name2text, name2wavpath)

    @property
    def info(self):
        pass

class Inputtr2017009(InputCollectBase):
    """
    Example: demo.wav *.txt
    000093-0000.wav 臭豆腐。        OK
    000093-0001.wav #NG     NN
    000093-0002.wav 你好。  OK
    000093-0003.wav #NG     NN
    000093-0004.wav #NG     NN
    000093-0005.wav 你好！  OK
    000093-0006.wav 你要吃什么？    OK
    """
    def __init__(self, data_dir, kformat_dir, mode="train", test_rate=0.001, lang="zh-CN"):
        self.data_dir = data_dir
        self.projname = os.path.basename(data_dir)
        self.kformat_dir = kformat_dir
        self.mode = mode
        self.test_rate = test_rate
        self.lang = lang

    def readTextFile(self, txt_files, name2wavpath):
        self.name2text = {}
        self.spk2name = defaultdict(list)
        self.name2spk = {}

        nums_perpart = len(txt_files) // 30 + 1
        parts = [txt_files[l:l+nums_perpart] for l in range(0, len(txt_files), nums_perpart)]
        threads = []
        for part in parts:
            t = threading.Thread(target=self.readTextPer, args=(part, name2wavpath,))
            threads.append(t)
            t.start()
        for t in threads:
            t.join()

        return self.name2text, self.spk2name, self.name2spk

    def readTextPer(self, txt_files, name2wavpath):
        for tii, txt_file in enumerate(txt_files):
            print(f"=> [{tii+1} / {len(txt_files)}]{txt_file}...")
            encoding_format = self.readEncodingType(txt_file)
            contents = open(txt_file, encoding=encoding_format).readlines()[1:]
            for line in contents:
                filename, text = line.strip().split("\t")[:2]
                text = re.sub("#[.*?]", "", text)
                if len(re.sub("[A-Z]{2}", "", text.replace(" ", ""))) == 0:
                    continue
                filename = filename[:-4]
                filename = self.projname + "_" + filename
                try:
                    filepath = name2wavpath[filename]
                except KeyError:
                    # print(filename)
                    continue
                try:
                    duration = self.readDuration(filepath)
                except Exception as e:
                    # print(e)
                    # print(filename)
                    continue
                text = re.sub("/[a-zA-Z]+:?-?(?:[a-zA-Z]+)?([0-9]+)?", " ", text)
                text = self.cleanText(text)
                if self.check(text):
                    # print("====================")
                    # print(text)
                    # print(self.check(text))
                    continue
                    # sys.exit()
                self.name2text[filename] = [[text, 0, duration]]
                self.spk2name[filename].append(filename)
                self.name2spk[filename] = filename

    def readWaveFiles(self, wav_files):
        name2wavpath = {}

        self.writeDurationThread(wav_files)
        # self.ReadMD5Thread(wav_files)
        for wav_file in tqdm(wav_files):
            filename = os.path.basename(wav_file)[:-4]
            filename = self.projname + "_" + filename
            name2wavpath[filename] = wav_file
        return name2wavpath

    def process(self):
        trtext_dir = os.path.join(self.data_dir, "product")
        trwave_dir = os.path.join(self.data_dir, "wave_org")
        print(trtext_dir)
        print(trwave_dir)
        txt_files = self.findOne(trtext_dir, "*.txt")
        print(f"Found text file: {len(txt_files)}...")
        wav_files = self.findOne(trwave_dir, "*.wav")
        print(f"Found wave file: {len(wav_files)}...")
        
        print("=> Processing wave files...")
        name2wavpath = self.readWaveFiles(wav_files)

        print("=> Processing text files...")
        name2text, spk2name, name2spk = self.readTextFile(txt_files, name2wavpath)
        self.splitAndWrite(spk2name, name2spk, name2text, name2wavpath)

    @property
    def info(self):
        pass

class InputAliTransFormatAdvanced(InputCollectBase):
    """
    For TR Department data from alibaba.
    one result.txt containing all wave's transcripts each ID
    example: ID0000/*.wav ID0000/result.txt
    there are a plenty of repeated files for a single batch, so we
    need to filter the old one out.[需要留下最新提交的版本的文本]
    """
    def __init__(self, data_dir, kformat_dir, mode="train", test_rate=0.001, text_index=1, name_index=0, lang="zh-CN"):
        self.data_dir = data_dir
        self.kformat_dir = kformat_dir
        self.mode = mode
        self.test_rate = test_rate
        self.projname = os.path.basename(data_dir)
        self.text_index = text_index
        self.name_index = name_index
        self.lang = lang
        self.info_dict = {}


    def cleanText(self, text):
        """
        Clean text to remove invailable tags or punctuations if needed!
        """
        text = re.sub("\[.*?\]", '', text)
        text = re.sub('<.*?>', ' ', text)
        text = re.sub('\(.*?\)', ' ', text)
        text = re.sub('[，。？,；、`！──－：‘’“”~《》．"；,/×—－─（）・／—"『〈〉』·「」<>.?!…;]', '', text)
        text = text.replace('*', ' ')
        text = " ".join(text.split())
        if "{" in text: # 有误读的 大括号里面写的还是正常的句子，这种读错的我们不要
            text = ""   # 例如 我 看 真正 的 安定团结 是 肯定 的 ， 现在 是 说 日 <B>{渐gai4}</B> 巩固
        text = self.strQ2B(text)
        return text

    def readTransFile(self, result_files):
        name2text = defaultdict(list)
        spk2name = defaultdict(list)
        name2spk = {}
        total_dur = 0
        for retfile in tqdm(result_files):
            encoding_format = self.readEncodingType(retfile)
            contents = open(retfile, encoding=encoding_format).readlines()
            if contents[0].split("\t")[self.text_index] not in ["文本", "转写结果"]:
                continue
            for line in contents[1:]:
                line = line.strip()
                if not line:
                    continue
                try:
                    name = line.strip().split("\t")[self.name_index]
                    text = line.strip().split("\t")[self.text_index]
                    time_info = line.strip().split("\t")[-1]
                except ValueError:
                    print(retfile)
                    print(line)
                    sys.exit()
                try:
                    start_time, end_time = re.findall("\[([0-9]+\.?(?:[0-9]+)?)\]", time_info)
                    start_time, end_time = float(start_time), float(end_time)
                except Exception:
                    print(line)
                    print(time_info)
                    # sys.exit()
                    continue
                name = self.projname + "_" + name[:-4]
                text = re.sub("（.*?）", " ", text)
                text = self.cleanText(text).replace("\\", "")
                text = text.replace("@", " at ")
                text = self.strQ2B(text)
                if self.projname in ["RS-TR-118_wangyi", "RS-TR-122_wangyi", "RS-TR-118", "RS-TR-122"] :
                    text = re.sub("n:[a-zA-Z]+(?:\-[a-zA-Z]+)? ", " ", text)
                    text = re.sub("\[.*?\]", " ", text)
                    text = re.sub("\(.*?\)", " ", text).replace("-/-", " ").replace("=", " ").replace("~", " ")
                else:
                    text = text.replace("[", " ").replace("]", " ")
                if self.check(text):
                    print(text)
                    # sys.exit()
                    continue
                total_dur += end_time - start_time
                name2text[name].append([text, start_time, end_time])
                name2spk[name] = name
                spk2name[name].append(name)
        print("Total Duration : ", total_dur)
        return name2text, spk2name, name2spk

    def filterOldOut(self, trtext_dir):
        result_file = []
        for batch in os.listdir(trtext_dir):
            batch = os.path.join(trtext_dir, batch)
            result_files = Path(batch).rglob("result.txt")
            result_files = [str(ele) for ele in result_files]
            if len(result_files) == 0:
                continue
            have_pass = False
            if len(result_files) == 1:
                result_file.append(result_files[0])
                have_pass = True
            name2time = {}
            for retfile in result_files:
                cons = open(retfile).readlines()[0].strip().split("\t")
                if "pass" in retfile:
                    have_pass = True
                    result_file.append(retfile)
                    break
                timeinfo = re.findall("[0-9]{8}", retfile)
                if timeinfo:
                    name2time[timeinfo[-1]] = retfile
                    have_pass = True
                    continue
                if "返工" in retfile:
                    result_file.append(retfile)
                    have_pass = True
                # print(os.path.basename(os.path.dirname(retfile)))
                # print(index)
            if not have_pass:
                print(batch)
                print(result_files)
            # print(result_files)
            if name2time:
                name2time_sorted = sorted(name2time.items(), key=lambda x:x[0], reverse=True)
                result_file.append(name2time_sorted[0][1])
        return result_file 
        
    def process(self):
        trtext_dir = os.path.join(self.data_dir, "product")
        trwave_dir = os.path.join(self.data_dir, "wave_org")
        
        result_file = self.filterOldOut(trtext_dir)

        wav_files = self.findOne(self.data_dir, "*.wav")

        name2text, spk2name, name2spk = self.readTransFile(result_file)

        name2wavpath = {}
        # self.ReadMD5Thread(wav_files)
        for wav_file in tqdm(wav_files):
            filename = os.path.basename(wav_file)[:-4]
            filename = self.projname + "_" + filename
            name2wavpath[filename] = wav_file
        
        self.splitAndWrite(spk2name, name2spk, name2text, name2wavpath)

    @property
    def info(self):
        return self.info_dict

class InputTransFormatLabel11thLong(InputCollectBase):
    """
    编号    文本    是否有效    身份    性别    口音    底噪    情绪    异常音  说话方式    时段
    LogATD-20200708-151841_002  哎我觉得。  有效    c1  [M] 无  无  正面    无  正常    [8.412][9.074]

    音频名  文本    性别    是否有口音  是否为儿童  发音边界
    8811.mp3.left.814.pcm.wav   【~】 呢很多人都说了小龙因为他在广州他团队在广州可能受到更独立更便于他这种自由的去  男  否  否  [0.406][6.697]
    """
    def __init__(self, data_dir, kformat_dir, mode="train", test_rate=0.001, lang="zh-CN"):
        self.data_dir = data_dir
        self.kformat_dir = kformat_dir
        self.mode = mode
        self.test_rate = test_rate
        self.projname = os.path.basename(data_dir)
        self.lang = lang
        self.info_dict = {}

    def readTransFile(self, result_files):
        name2text = defaultdict(list)
        spk2name = defaultdict(list)
        name2spk = {}
        total_dur = 0
        for retfile in tqdm(result_files):
            if "invalid" in retfile:
                continue
            encoding_format = self.readEncodingType(retfile)
            contents = open(retfile, encoding=encoding_format).readlines()
            if contents[0].split("\t")[0] not in ["编号", "音频名"]:
                print(f"HEAD NOT TRUE!{contents[0]}")
                continue
            filename = os.path.basename(retfile)[:-4]
            filename = self.projname + "_" + filename
            for line in contents[1:]:
                line = line.strip()
                if not line:
                    continue
                try:
                    _, text, flag, *_, time_info = line.strip().split('\t')
                except ValueError:
                    print(retfile)
                    print(line)
                    # sys.exit()
                    continue
                if self.projname != "TR2018068":
                    if flag != "有效":
                        continue    
                start_time, end_time = re.findall("\[([0-9]+\.?(?:[0-9]+)?)\]", time_info)
                start_time, end_time = float(start_time), float(end_time)
                text = self.cleanText(text)
                text = text.replace("_", " ")
                if self.check(text):
                    print(text)
                    continue
                total_dur += end_time - start_time
                name2text[filename].append([text, start_time, end_time])
            name2spk[filename] = filename
            spk2name[filename].append(filename)
        print("Total Duration : ", total_dur / 3600)
        return name2text, spk2name, name2spk

    def process(self):
        trtext_dir = os.path.join(self.data_dir, "product")
        trwave_dir = os.path.join(self.data_dir, "wave_org")
        if self.projname == "AliMeeting":
            trtext_dir = self.data_dir
            trwave_dir = self.data_dir

        text_files = self.findOne(trtext_dir, "*.txt")
        name2text, spk2name, name2spk = self.readTransFile(text_files)
        
        wav_files = self.findOne(trwave_dir, "*.wav")

        name2wavpath = {}
        # self.ReadMD5Thread(wav_files)
        for wav_file in tqdm(wav_files):
            filename = os.path.basename(wav_file)[:-4]
            filename = self.projname + "_" + filename
            name2wavpath[filename] = wav_file

        self.splitAndWrite(spk2name, name2spk, name2text, name2wavpath)

    @property
    def info(self):
        return self.info_dict

class InputTransFormatCSVLong(InputCollectBase):
    """
    编号,文本,是否有效,身份,性别,时段
    001,你好~ 致电一汽大众~ 四s店销售顾问申晴，很高兴为您服务，请问。,有效,客服1,女,[0][5.191]
    002,重叠音,无效,,,[5.191][6.941]
    003,你那石景山西路那边吗？,有效,客户1,男,[6.941][8.75]
    004,对。,有效,客服1,女,[8.75][9.322]
    """
    def __init__(self, data_dir, kformat_dir, mode="train", test_rate=0.001, lang="zh-CN"):
        self.data_dir = data_dir
        self.kformat_dir = kformat_dir
        self.mode = mode
        self.test_rate = test_rate
        self.projname = os.path.basename(data_dir)
        self.info_dict = {}
        self.lang = lang

    def readTransFile(self, result_files):
        name2text = defaultdict(list)
        spk2name = defaultdict(list)
        name2spk = {}
        total_dur = 0
        for retfile in tqdm(result_files):
            encoding_format = self.readEncodingType(retfile)
            contents = open(retfile, encoding=encoding_format).readlines()
            filename = os.path.basename(retfile)[:-4]
            filename = self.projname + "_" + filename
            for line in contents[1:]:
                line = line.strip()
                if not line:
                    continue
                try:
                    _, text, flag, *_, time_info = line.strip().split(',')
                except ValueError:
                    print(retfile)
                    print(line)
                    # sys.exit()
                    continue
                if flag != "有效":
                    continue    
                start_time, end_time = re.findall("\[([0-9]+\.?(?:[0-9]+)?)\]", time_info)
                start_time, end_time = float(start_time), float(end_time)
                text = self.cleanText(text)
                text = text.replace("_", " ")
                if self.check(text):
                    print(text)
                    continue
                total_dur += end_time - start_time
                name2text[filename].append([text, start_time, end_time])
            name2spk[filename] = filename
            spk2name[filename].append(filename)
        print("Total Duration : ", total_dur / 3600)
        return name2text, spk2name, name2spk

    def process(self):
        trtext_dir = os.path.join(self.data_dir, "product")
        trwave_dir = os.path.join(self.data_dir, "wave_org")
        if self.projname == "AliMeeting":
            trtext_dir = self.data_dir
            trwave_dir = self.data_dir

        text_files = self.findOne(trtext_dir, "*.CSV")
        name2text, spk2name, name2spk = self.readTransFile(text_files)
        
        wav_files = self.findOne(trwave_dir, "*.wav")

        name2wavpath = {}
        # self.ReadMD5Thread(wav_files)
        for wav_file in tqdm(wav_files):
            filename = os.path.basename(wav_file)[:-4]
            filename = self.projname + "_" + filename
            name2wavpath[filename] = wav_file

        self.splitAndWrite(spk2name, name2spk, name2text, name2wavpath)

    @property
    def info(self):
        return self.info_dict

class InputTransFormatDATLong(InputCollectBase):
    """
    # encoding=utf-8
        6.250   6.887   "<CUS/><FILLER>喂</FILLER>"
        6.887   9.555   "<OP/><FILLER>哎</FILLER>您好请问是<NAME>李辉</NAME><NAME>李</NAME>先生<FILLER>吗</FILLER>"
        9.555   10.806  "<CUS/><FILLER>嗯</FILLER>是你哪里<FILLER>呀</FILLER>"
    """
    def __init__(self, data_dir, kformat_dir, mode="train", test_rate=0.001, lang="zh-CN"):
        self.data_dir = data_dir
        self.kformat_dir = kformat_dir
        self.mode = mode
        self.test_rate = test_rate
        self.projname = os.path.basename(data_dir)
        self.lang = lang
        self.info_dict = {}

    def readTransFile(self, result_files):
        name2text = defaultdict(list)
        spk2name = defaultdict(list)
        name2spk = {}
        total_dur = 0
        for retfile in tqdm(result_files):
            if "invalid" in retfile:
                continue
            encoding_format = self.readEncodingType(retfile)
            contents = open(retfile, encoding=encoding_format).readlines()
            filename = os.path.basename(retfile)[:-4]
            filename = self.projname + "_" + filename
            for line in contents:
                line = line.strip()
                if line.startswith("#"):
                    continue
                try:
                    start_time, end_time, text = line.split('\t')
                except ValueError:
                    print(retfile)
                    print(line)
                    continue
                start_time, end_time = float(start_time), float(end_time)
                text = self.cleanText(text).replace("-", " ")
                text = re.sub("<.*?>", "", text)
                if self.check(text):
                    print(text)
                    continue
                total_dur += end_time - start_time
                name2text[filename].append([text, start_time, end_time])
            name2spk[filename] = filename
            spk2name[filename].append(filename)
        print("Total Duration : ", total_dur / 3600)
        return name2text, spk2name, name2spk

    def process(self):
        trtext_dir = os.path.join(self.data_dir, "product")
        trwave_dir = os.path.join(self.data_dir, "wave_org")

        text_files = self.findOne(trtext_dir, "*.dat")
        name2text, spk2name, name2spk = self.readTransFile(text_files)
        
        wav_files = self.findOne(trwave_dir, "*.wav")

        name2wavpath = {}
        # self.ReadMD5Thread(wav_files)
        for wav_file in tqdm(wav_files):
            filename = os.path.basename(wav_file)[:-4]
            filename = self.projname + "_" + filename
            name2wavpath[filename] = wav_file

        self.splitAndWrite(spk2name, name2spk, name2text, name2wavpath)

    @property
    def info(self):
        return self.info_dict

class InputRS9thTransFormat(InputCollectBase):
    """
    For TR Department data from RS format
    音频ID  音频名称    是否有效    转写结果    性别    是否有英语  是否有底噪  是否有口音  起止点静音
    """
    def __init__(self, data_dir, kformat_dir, name_index=1, text_index=3, mode="train", test_rate=0.001, lang="zh-CN"):
        self.data_dir = data_dir
        self.kformat_dir = kformat_dir
        self.mode = mode
        self.test_rate = test_rate
        self.projname = os.path.basename(data_dir)
        self.name_index = name_index
        self.text_index = text_index
        self.lang = lang
        self.info_dict = {}

    def readTransFile(self, result_files):
        name2text = defaultdict(list)
        spk2name = defaultdict(list)
        name2spk = {}
        for result_file in tqdm(result_files):
            if "invalid" in result_file:
                continue
            encoding_format = self.readEncodingType(result_file)
            try:
                contents = open(result_file, encoding=encoding_format).readlines()
            except UnicodeDecodeError:
                continue
            for line in contents[1:]:
                line = line.strip()
                try:
                    eles = line.strip().split('\t')
                    try:
                        name = eles[self.name_index]
                        text = eles[self.text_index]
                    except IndexError:
                        continue
                    time_info = eles[-1]
                except ValueError:
                    print(result_file)
                    print(line)
                    continue
                try:
                    start_time, end_time = re.findall("\[([0-9]+\.?(?:[0-9]+)?)\]", time_info)
                except ValueError:
                    print(result_file)
                    print(line)
                    continue
                start_time, end_time = float(start_time), float(end_time)
                if self.projname not in ["RS-TR-201"]:
                    name = self.projname + "_" + name[:-4]
                else:
                    name = self.projname + "_" + name
                text = self.cleanText(text).replace("〇", "零")
                text = re.sub("{.*?}", "", text)
                text = re.sub("\(.*?\]", "", text)
                text = re.sub("\[.*?\)\)?", "", text)
                text = re.sub("\(.*?\)", "", text)
                text = re.sub("{.*?\)", "", text)
                # text = re.sub("\)", "", text)
                if self.check(text):
                    print(text)
                    print(line)
                    continue
                name2text[name].append([text, start_time, end_time])
                name2spk[name] = name
                spk2name[name].append(name)
        return name2text, spk2name, name2spk

    def process(self):
        trtext_dir = os.path.join(self.data_dir, "product")
        trwave_dir = os.path.join(self.data_dir, "wave_org")
        print(f"=> Find text files....")
        result_files = self.findOne(trtext_dir, "*result*.txt")
        print(f"=> Find wave files....")
        wav_files = self.findOne(trwave_dir, "*.wav")
        print(f"Found text file: {len(result_files)}...")
        print(f"Found wave file: {len(wav_files)}...")
        print(f"=> Processing Text Files...")
        name2text, spk2name, name2spk = self.readTransFile(result_files)

        name2wavpath = {}
        print(f"=> Processing Wave Files...")
        # # self.ReadMD5Thread(wav_files)
        for wav_file in tqdm(wav_files):
            filename = os.path.basename(wav_file)[:-4]
            filename = self.projname + "_" + filename
            name2wavpath[filename] = wav_file
        
        self.splitAndWrite(spk2name, name2spk, name2text, name2wavpath)


    @property
    def info(self):
        return self.info_dict

class InputSTMTransFormat(InputCollectBase):
    """
    For TR Department data from RS format
    812605011730586_3 1 Customer 0.0 3.789 <有效,女> 完了他那个报警器它就响，响完了以后他就那个断气了。
    812605011730586_7 1 Customer 0.0 3.25 <有效,女> 现在我今天插怎么插他那个就那个那个都没有气。
    812605011730586_9 1 Customer 0.0 1.5 <有效,女> 燃气表哪个厂家的啊。
    812605011730586_10 1 Customer 0.0 1.579 <无效> 
    """
    def __init__(self, data_dir, kformat_dir, mode="train", test_rate=0.001, lang="zh-CN"):
        self.data_dir = data_dir
        self.kformat_dir = kformat_dir
        self.mode = mode
        self.test_rate = test_rate
        self.lang = lang
        self.projname = os.path.basename(data_dir)
        self.info_dict = {}

    def readTransFile(self, result_files):
        name2text = defaultdict(list)
        spk2name = defaultdict(list)
        name2spk = {}
        for result_file in tqdm(result_files):
            encoding_format = self.readEncodingType(result_file)
            contents = open(result_file, encoding=encoding_format).readlines()
            for line in contents:
                line = line.strip()
                if "<无效>" in line:
                    continue
                try:
                    name, _, _, start_time, end_time, vflag, *text = line.strip().split(' ')
                except ValueError:
                    print(result_file)
                    print(line)
                    continue
                if "有效" not in vflag:
                    print(line)
                    continue
                text = "".join(text)
                start_time, end_time = float(start_time), float(end_time)
                name = self.projname + "_" + name
                text = self.cleanText(text)
                if self.check(text):
                    print(text)
                    continue
                name2text[name].append([text, start_time, end_time])
                name2spk[name] = name
                spk2name[name].append(name)
        return name2text, spk2name, name2spk

    def process(self):
        trtext_dir = os.path.join(self.data_dir, "product")
        trwave_dir = os.path.join(self.data_dir, "wave_org")
        print(f"=> Find text files....")
        result_files = self.findOne(trtext_dir, "*.stm")
        print(f"=> Find wave files....")
        wav_files = self.findOne(trwave_dir, "*.wav")
        print(f"Found text file: {len(result_files)}...")
        print(f"Found wave file: {len(wav_files)}...")
        print(f"=> Processing Text Files...")
        name2text, spk2name, name2spk = self.readTransFile(result_files)

        name2wavpath = {}
        print(f"=> Processing Wave Files...")
        # self.ReadMD5Thread(wav_files)
        for wav_file in tqdm(wav_files):
            filename = os.path.basename(wav_file)[:-4]
            filename = self.projname + "_" + filename
            name2wavpath[filename] = wav_file
        
        self.splitAndWrite(spk2name, name2spk, name2text, name2wavpath)


    @property
    def info(self):
        return self.info_dict

class InputShortWithTimeInfoTransFormat(InputCollectBase):
    """
    For TR Department data from RS format
    音频名  文本    性别    是否有口音      是否为儿童      发音边界 
    2b8c7772111b4ed486f5b75f886cb8e3.raw.wav        太阳的后裔      男      否      否      [1.517][2.665]
    7cd23ad4af5a450ea60d1c04bf863e83.raw.wav        现在几点了      女      是      否      [0.697][2.014]
    """
    def __init__(self, data_dir, kformat_dir, mode="train", test_rate=0.001, lang='zh-CN'):
        self.data_dir = data_dir
        self.kformat_dir = kformat_dir
        self.mode = mode
        self.test_rate = test_rate
        self.projname = os.path.basename(data_dir)
        self.lang = lang
        self.info_dict = {}

    def readTransFile(self, result_files):
        name2text = defaultdict(list)
        spk2name = defaultdict(list)
        name2spk = {}
        for result_file in tqdm(result_files):
            if "invalid" in result_file:
                continue
            try:
                encoding_format = self.readEncodingType(result_file)
                contents = open(result_file, encoding=encoding_format).readlines()
            except Exception:
                continue
            for line in contents[1:]:
                line = line.strip()
                try:
                     name, text, *_, time_info = line.strip().split('\t')
                except ValueError:
                    print(result_file)
                    print(line)
                    continue
                try:
                    time_infos = re.findall("\[([0-9]+\.?(?:[0-9]+)?)\]", time_info)
                    if len(time_infos) == 2:
                        start_time, end_time = time_infos
                    elif len(time_infos) == 4:
                        if len(time_infos[0]) == 0:
                            start_time, end_time = time_infos[-2:]
                        else:
                            start_time, *_, end_time = time_infos
                    else:
                        print(line)
                        print(result_file)
                        continue
                except ValueError:
                    print(result_file)
                    print(line)
                    continue
                start_time, end_time = float(start_time), float(end_time)
                name = self.projname + "_" + name[:-4]
                text = text.replace("_", " ").replace("@", " at ")
                text = self.cleanText(text)
                if self.check(text):
                    print("----------------")
                    print(text)
                    continue
                name2text[name].append([text, start_time, end_time])
                name2spk[name] = name
                spk2name[name].append(name)
        return name2text, spk2name, name2spk

    def process(self):
        trtext_dir = os.path.join(self.data_dir, "product")
        trwave_dir = os.path.join(self.data_dir, "wave_org")
        if self.projname == "TR2017083":
            trwave_dir = os.path.join(self.data_dir, "product")
        print(f"=> Find text files....")
        result_files = self.findOne(trtext_dir, "*result*.txt")
        print(f"=> Find wave files....")
        wav_files = self.findOne(trwave_dir, "*.wav")
        print(f"Found text file: {len(result_files)}...")
        print(f"Found wave file: {len(wav_files)}...")
        print(f"=> Processing Text Files...")
        name2text, spk2name, name2spk = self.readTransFile(result_files)
        
        name2wavpath = {}
        print(f"=> Processing Wave Files...")
        # self.ReadMD5Thread(wav_files)
        for wav_file in tqdm(wav_files):
            filename = os.path.basename(wav_file)[:-4]
            filename = self.projname + "_" + filename
            name2wavpath[filename] = wav_file
        
        self.splitAndWrite(spk2name, name2spk, name2text, name2wavpath)

    @property
    def info(self):
        return self.info_dict

class InputQQ3thTransFormat(InputCollectBase):
    """
    TR2021036
    音频名称        时间    内容
    qq_trans_20201207_100108_166e714116d884703e2d85bfe1aff3649.wav  [0.215][3.551]  现在交钱吧，有位移一切的奢求吧，有解控。
    """
    def __init__(self, data_dir, kformat_dir, mode="train", test_rate=0.001, lang="zh-CN"):
        self.data_dir = data_dir
        self.kformat_dir = kformat_dir
        self.mode = mode
        self.test_rate = test_rate
        self.projname = os.path.basename(data_dir)
        self.info_dict = {}
        self.lang = lang

    def readTransFile(self, result_files):
        name2text = defaultdict(list)
        spk2name = defaultdict(list)
        name2spk = {}
        for result_file in result_files:
            encoding_format = self.readEncodingType(result_file)
            contents = open(result_file, encoding=encoding_format).readlines()
            for line in tqdm(contents[1:]):
                line = line.strip()
                try:
                    name, time_info, text = line.strip().split('\t')
                except ValueError:
                    print(result_file)
                    print(line)
                try:
                    start_time, end_time = re.findall("\[([0-9]+\.?(?:[0-9]+)?)\]", time_info)
                except ValueError:
                    print(result_file)
                    print(line)
                    continue
                start_time, end_time = float(start_time), float(end_time)
                name = self.projname + "_" + name[:-4]
                text = self.cleanText(text)
                if self.check(text):
                    print(text)
                    continue
                name2text[name].append([text, start_time, end_time])
                name2spk[name] = name
                spk2name[name].append(name)
        return name2text, spk2name, name2spk

    def process(self):
        trtext_dir = os.path.join(self.data_dir, "product")
        trwave_dir = os.path.join(self.data_dir, "wave_org")
        print(f"=> Find text files....")
        result_files = self.findOne(trtext_dir, "result.txt")
        print(f"=> Find wave files....")
        wav_files = self.findOne(trwave_dir, "*.wav")
        print(f"Found text file: {len(result_files)}...")
        print(f"Found wave file: {len(wav_files)}...")
        print(f"=> Processing Text Files...")
        name2text, spk2name, name2spk = self.readTransFile(result_files)
        
        name2wavpath = {}
        print(f"=> Processing Wave Files...")
        # self.ReadMD5Thread(wav_files)
        for wav_file in tqdm(wav_files):
            filename = os.path.basename(wav_file)[:-4]
            filename = self.projname + "_" + filename
            name2wavpath[filename] = wav_file
        
        self.splitAndWrite(spk2name, name2spk, name2text, name2wavpath)

    @property
    def info(self):
        return self.info_dict

class InputTRTextGridSubFolder(InputCollectBase):
    """
    转写的Textgrid格式 [TR2020024]
    """
    def __init__(self, data_dir, kformat_dir, mode="train", test_rate=0.001, item_index=0, lang='zh-CN'):
        self.data_dir = data_dir
        self.kformat_dir = kformat_dir
        self.mode = mode
        self.test_rate = test_rate
        self.projname = os.path.basename(data_dir)
        self.info_dict = {}
        self.item_index = item_index
        self.lang = lang

    def readTextGridFile(self, batch_folders, trtext_dir):
        name2text = defaultdict(list)
        spk2name = defaultdict(list)
        name2spk = {}
        duration_total = 0
        for batch in batch_folders:
            print(f"=> {batch}...")
            subtext_dir = os.path.join(trtext_dir, batch)
            tg_files = self.findOne(subtext_dir, "*.TextGrid")
            for tg_file in tqdm(tg_files):
                filename = os.path.basename(tg_file)[:-9]
                if filename[-1] == "-":
                    filename = filename[:-1]
                filename = self.projname + "_" + batch + "--" + filename
                tg = TextGrid()
                try:
                    tg.read(tg_file)
                except Exception as e:
                    print("ERROR WITH TEXTGRID FILE")
                    print(tg_file)
                    print(e)
                    continue
                try:
                    for info in tg.tiers:
                        if info.name == "PEOPLE" or info.name == "text":
                            break
                    if info.name != "PEOPLE" and info.name != "text":
                        info = tg.tiers[self.item_index]
                    # if info.name == "SPEAKER":
                    #     info = tg.tiers[1] ## BUG
                except Exception:
                    print("ERROR WITH TIERS")
                    info = tg.tiers[0]
                for ele in info:
                    text, maxt, mint = ele.mark, float(ele.maxTime), float(ele.minTime)
                    if text == "sil" or text == "noi":
                        continue
                    text = self.cleanText(text).replace("@", " at ")
                    if self.check(text):
                        try:
                            print("Error :", text, tg_file)
                        except Exception:
                            continue
                        print(f"------\n{self.check(text)}\n-------")
                        continue
                    if len(text) == 0:
                        continue
                    duration_total += maxt - mint
                    name2text[filename].append([text, mint, maxt])
                name2spk[filename] = filename
                spk2name[filename].append(filename)
        print("Total Duration[h]: ", duration_total / 3600)
        return name2text, spk2name, name2spk


    def process(self):
        trtext_dir = os.path.join(self.data_dir, "product")
        trwave_dir = os.path.join(self.data_dir, "wave_org")
        batch_folders = os.listdir(trtext_dir)
        name2text, spk2name, name2spk = self.readTextGridFile(batch_folders, trtext_dir)
        
        name2wavpath = {}
        for batch in batch_folders:
            print(f"=> {batch}...")
            subwav_folder = os.path.join(trwave_dir, batch)
            wav_files = self.findOne(subwav_folder, "*.wav")
            # self.ReadMD5Thread(wav_files)
            for wav_file in tqdm(wav_files):
                filename = os.path.basename(wav_file)[:-4]
                filename = self.projname + "_" + batch + "--" + filename
                name2wavpath[filename] = wav_file

        self.splitAndWrite(spk2name, name2spk, name2text, name2wavpath)

class InputTR2021106(InputCollectBase):
    """
    [00:00:00,00:00:00.761]  [Spk_noise] [none,none]
    [00:00:00.761,00:00:12.759] [Spk] [A,male]  我们可能是分成这几个角度啊，就第一个的话呢就是合同呃从合同金额的角度，我们其实至少是在那个目标
    上我们定的还是相对来讲比较积极的。我们可能是要再。
    """
    def __init__(self, data_dir, kformat_dir, mode="train", test_rate=0.001, lang="zh-CN"):
        self.data_dir = data_dir
        self.kformat_dir = kformat_dir
        self.mode = mode
        self.test_rate = test_rate
        self.lang = lang
        self.projname = os.path.basename(data_dir)
        self.info_dict = {}

    def timeTranslator(self, ftime):
        h, m, s = ftime.split(":")
        h, m, s = float(h), float(m), float(s)
        duration = h * 3600 + m * 60 + s 
        return duration

    def readTextFile(self, text_files):
        name2text = defaultdict(list)
        spk2name = defaultdict(list)
        name2spk = {}
        duration_total = 0
        textnames = defaultdict(list)
        for text_file in tqdm(text_files):
            filename = os.path.basename(text_file)[:-4]
            filename = self.projname + "_" + filename
            textnames[filename].append(text_file)
            encoding_format = self.readEncodingType(text_file)
            cons = open(text_file, encoding=encoding_format, errors="ignore").readlines()
            for line in cons:
                line = line.strip()
                if not line.startswith("["):
                    continue
                text = line.split("]", 1)[1]
                text = re.sub("\[.*?\]", " ", text)
                text = self.cleanText(text).replace("Sil", " ")
                text = re.sub("Spk_[a-zA-Z]+", " ", text).replace("Noise", " ")
                if not text.replace(" ", ""):
                    continue
                if self.check(text):
                    print(line)
                    print(text)
                    # sys.exit()
                    continue
                # [00:00:04.527,00:00:06.383,A]
                start_time, end_time = re.findall("\[(.*?)\]", line)[0].split(",")[:2]
                start_time, end_time = self.timeTranslator(start_time), self.timeTranslator(end_time)
                if end_time <= start_time:
                    print(text_file, line)
                    continue
                name2text[filename].append([text, start_time, end_time])
                duration_total += end_time - start_time
            spk2name[filename].append(filename)
            name2spk[filename] = filename
        print("Total Duration: ", duration_total / 3600)
        return name2text, spk2name, name2spk


    def process(self):
        trtext_dir = os.path.join(self.data_dir, "product")
        trwave_dir = os.path.join(self.data_dir, "wave_org")

        text_files = self.findOne(trtext_dir, "*.txt")
        name2text, spk2name, name2spk = self.readTextFile(text_files)

        wav_files = self.findOne(trwave_dir, "*.wav")
        
        self.info_dict["num_textgrid_files"] = len(text_files)
        self.info_dict["num_wave_files"] = len(wav_files)

        name2wavpath = {}
        # self.ReadMD5Thread(wav_files)
        for wav_file in wav_files:
            filename = self.projname + "_" + os.path.basename(wav_file)[:-4]
            name2wavpath[filename] = wav_file

        self.splitAndWrite(spk2name, name2spk, name2text, name2wavpath)

class InputRSDPTransFormat(InputCollectBase):
    """
    For TR Department data from RS format
    [1.802] [6.034] 这几天一直很想，非常想。
    """
    def __init__(self, data_dir, kformat_dir, mode="train", test_rate=0.001, lang="zh-CN"):
        self.data_dir = data_dir
        self.kformat_dir = kformat_dir
        self.mode = mode
        self.test_rate = test_rate
        self.lang = lang
        self.projname = os.path.basename(data_dir)
        self.info_dict = {}

    def readTransFile(self, text_files):
        name2text = defaultdict(list)
        spk2name = defaultdict(list)
        name2spk = {}
        total_dur = 0
        for text_file in tqdm(text_files):
            if "属性" in text_file:
                continue
            try:
                encoding_format = self.readEncodingType(text_file)
                contents = open(text_file, encoding=encoding_format).readlines()
            except Exception:
                continue
            name = os.path.basename(text_file)[:-8]
            name = self.projname + "_" + name
            for line in contents:
                line = line.strip()
                try:
                     start_time, end_time, text = line.strip().split('\t')
                except ValueError:
                    print(text_file)
                    print(line)
                    continue
                try:
                    start_time = re.findall("\[([0-9]+\.?(?:[0-9]+)?)\]", start_time)[0]
                    end_time = re.findall("\[([0-9]+\.?(?:[0-9]+)?)\]", end_time)[0]
                    start_time, end_time = float(start_time), float(end_time)
                except Exception:
                    print(line)
                    continue
                total_dur += end_time - start_time
                text = text.replace("_", " ").replace("@", " at ")
                text = re.sub("\\.*?", " ", text)
                text = text.replace("\\ n : b r e a t h", " ")
                text = text.replace("\\ n : l i p s m a c k", " ")
                text = text.replace("\\ n : l a u g h t e r", " ")
                text = text.replace("\\ n : c o u g h", " ")
                text = text.replace("\\ n : s i g h", " ")
                text = text.replace("\\ n : t h r o a t - c l e a r", " ")
                text = text.replace("\\ n p s", " ")
                text = text.replace("\\ n", " ")
                text = text.replace("\\ s p h", " ")
                text = text.replace("~", "").replace(":", "")
                text = self.cleanText(text)
                if self.check(text):
                    print("----------------")
                    print(text_file)
                    print(text)
                    continue
                name2text[name].append([text, start_time, end_time])
            name2spk[name] = name
            spk2name[name].append(name)
        print("Total Duration: ", total_dur / 3600)
        return name2text, spk2name, name2spk

    def process(self):
        trtext_dir = os.path.join(self.data_dir, "product")
        trwave_dir = os.path.join(self.data_dir, "wave_org")
        if self.projname == "TR2017083":
            trwave_dir = os.path.join(self.data_dir, "product")
        print(f"=> Find text files....")
        text_files = self.findOne(trtext_dir, "*.txt")
        print(f"=> Find wave files....")
        wav_files = self.findOne(trwave_dir, "*.wav")
        print(f"Found text file: {len(text_files)}...")
        print(f"Found wave file: {len(wav_files)}...")
        print(f"=> Processing Text Files...")
        name2text, spk2name, name2spk = self.readTransFile(text_files)
        
        name2wavpath = {}
        print(f"=> Processing Wave Files...")
        # self.ReadMD5Thread(wav_files)
        for wav_file in tqdm(wav_files):
            filename = os.path.basename(wav_file)[:-4]
            filename = self.projname + "_" + filename
            name2wavpath[filename] = wav_file
        
        self.splitAndWrite(spk2name, name2spk, name2text, name2wavpath)

    @property
    def info(self):
        return self.info_dict


class InputTencentKpl(InputCollectBase):
    """
    [00:00:00,00:00:00.761]  [Spk_noise] [none,none]
    [00:00:00.761,00:00:12.759] [Spk] [A,male]  我们可能是分成这几个角度啊，就第一个的话呢就是合同呃从合同金额的角度，我们其实至少是在那个目标
    上我们定的还是相对来讲比较积极的。我们可能是要再。
    """
    def __init__(self, data_dir, kformat_dir, mode="train", test_rate=0.001, lang="zh-CN"):
        self.data_dir = data_dir
        self.kformat_dir = kformat_dir
        self.mode = mode
        self.test_rate = test_rate
        self.lang = lang
        self.projname = os.path.basename(data_dir)
        self.info_dict = {}

    def timeTranslator(self, ftime):
        h, m, s = ftime.split(":")
        h, m, s = float(h), float(m), float(s)
        duration = h * 3600 + m * 60 + s 
        return duration

    def readTextFile(self, text_files):
        name2text = defaultdict(list)
        spk2name = defaultdict(list)
        name2spk = {}
        duration_total = 0
        textnames = defaultdict(list)
        for text_file in tqdm(text_files):
            filename = os.path.basename(text_file)[:-4]
            filename = self.projname + "_" + filename
            textnames[filename].append(text_file)
            encoding_format = self.readEncodingType(text_file)
            cons = open(text_file, encoding=encoding_format, errors="ignore").readlines()
            for line in cons:
                line = line.strip()
                if not line.startswith("["):
                    continue
                if len(line.split("\t")) != 4:
                    continue
                time_info, *_, text = line.split("\t")
                text = self.cleanText(text)
                if not text.replace(" ", ""):
                    continue
                if self.check(text):
                    print(line)
                    print(text)
                    continue
                # [00:00:04.527,00:00:06.383,A]
                try:
                    start_time, end_time = re.findall("\[(.*?)\]", line)[0].split("-")
                except Exception:
                    start_time, end_time = re.findall("\[(.*?)\]", line)[0].split(",")
                start_time, end_time = float(start_time), float(end_time)
                if end_time <= start_time:
                    print(text_file, line)
                    continue
                name2text[filename].append([text, start_time, end_time])
                duration_total += end_time - start_time
            spk2name[filename].append(filename)
            name2spk[filename] = filename
        print("Total Duration: ", duration_total / 3600)
        return name2text, spk2name, name2spk


    def process(self):
        print("=> Task starting...")
        print(f"=> {self.data_dir}")
        text_files = self.findOne(self.data_dir, "*.txt")
        name2text, spk2name, name2spk = self.readTextFile(text_files)

        wav_files = self.findOne(self.data_dir, "*.wav")
        
        self.info_dict["num_textgrid_files"] = len(text_files)
        self.info_dict["num_wave_files"] = len(wav_files)

        name2wavpath = {}
        # self.ReadMD5Thread(wav_files)
        for wav_file in wav_files:
            filename = self.projname + "_" + os.path.basename(wav_file)[:-4]
            name2wavpath[filename] = wav_file

        self.splitAndWrite(spk2name, name2spk, name2text, name2wavpath)


class InputBytedanceJAJP(InputCollectBase):
    """
    """
    def __init__(self, data_dir, kformat_dir, mode="train", test_rate=0.001, lang="ja-JP"):
        self.data_dir = data_dir
        self.kformat_dir = kformat_dir
        self.mode = mode
        self.test_rate = test_rate
        self.info_dict = {}
        self.lang = lang

    def check(self, text):
        """
        Check if there are invalid characters.
        """
        no_char = re.sub(r"[\u4E00-\u9FFF]", lambda x:"", text) # 匹配中文简体和繁体：
        no_char = re.sub(r"[\u3040-\u309f]", lambda x:"", no_char) # 匹配日文平假名
        no_char = re.sub(r"[\u30a0-\u30ff]", lambda x:"", no_char) # 匹配日文片假名
        no_char = re.sub(r"[a-zA-Z]+", "", no_char)
        no_char = re.sub(f"'", "", no_char).replace("-", "")
        no_char = re.sub("[，。？、！]", '', no_char).replace(" ", "")
        no_char = re.sub("[グ〆〒﨑グ〇々]", "", no_char)
        return no_char

    def readTextFile(self, text_files):
        name2text = defaultdict(list)
        spk2name = defaultdict(list)
        name2spk = {}
        duration_total = 0
        for text_file in tqdm(text_files):
            encoding_format = self.readEncodingType(text_file)
            cons = open(text_file, encoding=encoding_format, errors="ignore").readlines()
            for line in cons:
                line = line.strip()
                name, start_time, end_time, text = line.split("\t")
                name = name[:-4]
                start_time, end_time = float(start_time), float(end_time)
                text = self.cleanText(text)
                if not text.replace(" ", ""):
                    continue
                if self.check(text):
                    if not re.findall("[0-9]", text) and self.check(text) != "ø":
                        print(self.check(text))
                        print(line)
                        print(text)
                    continue
                name2text[name].append([text, start_time, end_time])
                duration_total += end_time - start_time
                spk2name[name] = [name]
                name2spk[name] = name
        print("Total Duration: ", duration_total / 3600)
        return name2text, spk2name, name2spk

    def process(self):

        text_files = self.findOne(self.data_dir, "*.txt")
        name2text, spk2name, name2spk = self.readTextFile(text_files)

        wav_files = self.findOne(self.data_dir, "*.wav")

        name2wavpath = {}
        for wav_file in wav_files:
            filename = os.path.basename(wav_file)[:-4]
            name2wavpath[filename] = wav_file

        self.splitAndWrite(spk2name, name2spk, name2text, name2wavpath)

class InputBytedanceKOKR(InputCollectBase):
    """
    """
    def __init__(self, data_dir, kformat_dir, mode="train", test_rate=0.001):
        self.data_dir = data_dir
        self.mode = mode
        self.test_rate = test_rate
        self.info_dict = {}
        self.kformat_dir = kformat_dir
        self.projname = os.path.basename(data_dir)

    def check(self, text):
        """
        Check if there are invalid characters.
        """
        no_char = re.sub(r"[\uac00-\ud7ff]", lambda x:"", text) # 3130-318F
        no_char = re.sub(r"[a-zA-Z]+", "", no_char)
        no_char = re.sub(f"'", "", no_char).replace("-", "")
        no_char = re.sub("[，。？、！]", '', no_char).replace(" ", "")
        no_char = re.sub("[ㅋㅎㅠ죠ㅜㅈㅣㅈㅇㄱㅂㄷㅂㄹㅅ]", "", no_char)
        return no_char

    def readTextFile(self, text_files):
        name2text = defaultdict(list)
        spk2name = defaultdict(list)
        name2spk = {}
        duration_total = 0
        for text_file in tqdm(text_files):
            encoding_format = self.readEncodingType(text_file)
            cons = open(text_file, encoding=encoding_format, errors="ignore").readlines()
            for line in cons:
                line = line.strip()
                name, start_time, end_time, text = line.split("\t")
                name = name[:-4]
                start_time, end_time = float(start_time), float(end_time)
                text = self.cleanText(text)
                if not text.replace(" ", ""):
                    continue
                # if self.check(text):
                #     print(line)
                #     print(text)
                #     print(self.check(text))
                #     # sys.exit()
                #     continue
                name2text[name].append([text, start_time, end_time])
                duration_total += end_time - start_time
                spk2name[name] = [name]
                name2spk[name] = name
        print("Total Duration: ", duration_total / 3600)
        return name2text, spk2name, name2spk

    def process(self):

        text_files = self.findOne(self.data_dir, "*.txt")
        name2text, spk2name, name2spk = self.readTextFile(text_files)

        wav_files = self.findOne(self.data_dir, "*.wav")

        name2wavpath = {}
        for wav_file in wav_files:
            filename = os.path.basename(wav_file)[:-4]
            name2wavpath[filename] = wav_file

        self.splitAndWrite(spk2name, name2spk, name2text, name2wavpath)

class InputBytedancePTBR(InputCollectBase):
    """
    """
    def __init__(self, data_dir, kformat_dir, mode="train", test_rate=0.001):
        self.data_dir = data_dir
        self.kformat_dir = kformat_dir
        self.mode = mode
        self.test_rate = test_rate
        self.info_dict = {}

    def readTextFile(self, text_files):
        name2text = defaultdict(list)
        spk2name = defaultdict(list)
        name2spk = {}
        duration_total = 0
        for text_file in tqdm(text_files):
            encoding_format = self.readEncodingType(text_file)
            cons = open(text_file, encoding=encoding_format, errors="ignore").readlines()
            for line in cons:
                line = line.strip()
                name, start_time, end_time, text = line.split("\t")
                name = name[:-4]
                start_time, end_time = float(start_time), float(end_time)
                text = self.cleanText(text)
                if not text.replace(" ", ""):
                    continue
                name2text[name].append([text, start_time, end_time])
                duration_total += end_time - start_time
                spk2name[name] = [name]
                name2spk[name] = name
        print("Total Duration: ", duration_total / 3600)
        return name2text, spk2name, name2spk

    def process(self):

        text_files = self.findOne(self.data_dir, "*.txt")
        name2text, spk2name, name2spk = self.readTextFile(text_files)

        wav_files = self.findOne(self.data_dir, "*.wav")

        name2wavpath = {}
        for wav_file in wav_files:
            filename = os.path.basename(wav_file)[:-4]
            name2wavpath[filename] = wav_file

        self.splitAndWrite(spk2name, name2spk, name2text, name2wavpath)


class InputUSTCFormat(InputCollectBase):
    """
    For TR Department data from alibaba.
    one result.txt containing all wave's transcripts each ID
    example: ID0000/*.wav ID0000/result.txt
    """
    def __init__(self, data_dir, kformat_dir, mode="train", test_rate=0.001, lang=""):
        self.data_dir = data_dir
        self.kformat_dir = kformat_dir
        self.mode = mode
        self.test_rate = test_rate
        self.projname = os.path.basename(data_dir)
        self.info_dict = {}
        self.lang = lang

    def readTransFile(self, info_files, name2seg):
        name2text = defaultdict(list)
        spk2name = defaultdict(list)
        name2spk = {}
        for info_file in info_files:
            print(f"processing {info_file}.")
            filedir = os.path.dirname(info_file)
            cons = open(info_file).readlines()
            for ii in range(0, len(cons), 4):
                filename = cons[ii].strip()[:-4] ## Azerbaijan_ekIsLg_F_29_0530.wav
                filepath = os.path.join(filedir, filename+".wav")
                spk = self.projname + "_" + "_".join(filename.split('_')[:-1])
                if not os.path.exists(filepath):
                    filename = filename.replace("Female", "F")
                    filename = filename.replace("Male", "M")
                    filepath = os.path.join(filedir, filename)
                if not os.path.exists(filepath):
                    print(f"{filepath} not existed!")
                filename = self.projname + "_" + filename
                text = cons[ii+1].strip()
                text = self.cleanText(text)
                spk2name[spk].append(filename)
                name2spk[filename] = spk
                start_time, end_time = name2seg[filename]
                name2text[filename] = [[text, start_time, end_time]]
        return name2text, spk2name, name2spk

    def process(self):
        txt_files = self.findOne(self.data_dir, "*.txt")
        wav_files = self.findOne(self.data_dir, "*.wav")


        name2wavpath = {}
        # self.writeDurationThread(wav_files)
        name2seg = {}
        for wav_file in tqdm(wav_files):
            filename = os.path.basename(wav_file)[:-4]
            filename = self.projname + "_" + filename
            name2wavpath[filename] = wav_file
            duration = self.readDuration(wav_file)
            name2seg[filename] = [0, duration]

        name2text, spk2name, name2spk = self.readTransFile(txt_files, name2seg)
        
        self.splitAndWrite(spk2name, name2spk, name2text, name2wavpath)


    @property
    def info(self):
        return self.info_dict



class InputTTSFormat(InputCollectBase):
    """
    For TR Department data from alibaba.
    one result.txt containing all wave's transcripts each ID
    example: ID0000/*.wav ID0000/result.txt
    """
    def __init__(self, data_dir, kformat_dir, mode="train", test_rate=0.001, lang=""):
        self.data_dir = data_dir
        self.kformat_dir = kformat_dir
        self.mode = mode
        self.test_rate = test_rate
        self.projname = os.path.basename(data_dir)
        self.info_dict = {}
        self.lang = lang

    def readTransFile(self, info_files, name2seg):
        name2text = defaultdict(list)
        spk2name = defaultdict(list)
        name2spk = {}
        for info_file in info_files:
            print(f"processing {info_file}.")
            cons = open(info_file, encoding='utf-8-sig').readlines()
            for ii in range(len(cons)):
                if not cons[ii].strip():
                    continue
                filename, text = cons[ii].strip().split("\t")
                spk = self.projname + "_" + filename
                filename = self.projname + "_" + filename
                text = self.cleanText(text)
                spk2name[spk].append(filename)
                name2spk[filename] = spk
                start_time, end_time = name2seg[filename]
                name2text[filename] = [[text, start_time, end_time]]
        return name2text, spk2name, name2spk

    def process(self):
        txt_files = self.findOne(self.data_dir, "*.txt")
        wav_files = self.findOne(self.data_dir, "*.wav")

        name2wavpath = {}
        self.writeDurationThread(wav_files)
        name2seg = {}
        for wav_file in tqdm(wav_files):
            filename = os.path.basename(wav_file)[:-4]
            filename = self.projname + "_" + filename
            name2wavpath[filename] = wav_file
            duration = self.readDuration(wav_file)
            name2seg[filename] = [0, duration]

        name2text, spk2name, name2spk = self.readTransFile(txt_files, name2seg)
        
        self.splitAndWrite(spk2name, name2spk, name2text, name2wavpath)

    @property
    def info(self):
        return self.info_dict

class ParlaSpeechFormat(InputCollectBase):
    """
    循环智能的数据
    Example: ParlaSpeech-HR.v1.0.jsonl.json  name.flac
    {'path': 'seg.qNpeHxO0WzA_1967.6-1986.11.flac', 'orig_file': '24 1 2020 - 16. sjednica, 9. saziv [qNpeHxO0WzA].wav', 'start': 1967.6, 'end': 1986.11, 'words': ['primjeni', 'željezničkih', 'paketa', 'kojima', 'se', 'postupno', 'i', 'nezaustavljivo', 'stvara', 'jedinstveni', 'europski', 'željeznički', 'prostor.', 'Provedba', 'ovog', 'zakona', 'nema', 'financijski', 'utjecaj', 'na', 'državni', 'proračun.', 'Hvala', 'lijepo.', 'Zahvaljujem', 'ministru.', 'Imamo', '7', 'replika,', 'prvi', 'je', 'kolega', 'Kirin.', 'Izvolite.'], 'word_start_times': [0, 0.51, 0.99, 1.5, 1.82, 1.99, 2.99, 3.12, 3.98, 4.4, 5.05, 5.53, 6.11, 7.01, 7.52, 7.76, 8.59, 9.0, 9.61, 10.26, 10.41, 10.81, 11.39, 11.57, 11.77, 12.24, 12.83, 13.27, 13.65, 14.09, 14.36, 14.43, 14.8, 15.23], 'norm_words': ['primjeni', 'željezničkih', 'paketa', 'kojima', 'se', 'postupno', 'i', 'nezaustavljivo', 'stvara', 'jedinstveni', 'europski', 'željeznički', 'prostor', 'provedba', 'ovog', 'zakona', 'nema', 'financijski', 'utjecaj', 'na', 'državni', 'proračun', 'hvala', 'lijepo', 'zahvaljujem', 'ministru', 'imamo', 'sedam', 'replika', 'prvi', 'je', 'kolega', 'kirin', 'izvolite'], 'norm_words_start_times': [0, 0.51, 0.99, 1.5, 1.82, 1.99, 2.99, 3.12, 3.98, 4.4, 5.05, 5.53, 6.11, 7.01, 7.52, 7.76, 8.59, 9.0, 9.61, 10.26, 10.41, 10.81, 11.39, 11.57, 11.77, 12.24, 12.83, 13.27, 13.65, 14.09, 14.36, 14.43, 14.8, 15.23], 'utterance_id_start': 'ParlaMint-HR_S16.u1923', 'utterance_id_end': 'ParlaMint-HR_S16.u1924', 'speaker_info': None, 'split': None}
    """
    def __init__(self, data_dir, kformat_dir, mode="test", test_rate=0.001):
        self.data_dir = data_dir
        self.kformat_dir = kformat_dir
        self.mode = mode
        self.test_rate = test_rate
        self.projname = "ParlaSpeech"
        self.info_dict = {}


    def readJsonFile(self, json_file):
        name2text = defaultdict(list)
        spk2name = defaultdict(list)
        name2spk = {}
        name2wavpath = {}
        print(json_file)
        contents = open(json_file).readlines()
        for line in tqdm(contents):
            line = line.strip()
            info = json.loads(line)
            name = info['path']
            wavpath = os.path.join(self.data_dir, 'hr_hr', name)
            duration = info['end'] - info['start']
            text = " ".join(info["norm_words"])
            name = self.projname + "_" + name[:-4]
            name2text[name].append([text, 0, duration])
            name2spk[name] = name
            spk2name[name].append(name)
            name2wavpath[name] = wavpath
        return name2text, spk2name, name2spk, name2wavpath

    def process(self):
        json_file = os.path.join(self.data_dir, "ParlaSpeech-HR.v1.0.jsonl")
        name2text, spk2name, name2spk, name2wavpath = self.readJsonFile(json_file)
        self.splitAndWrite(spk2name, name2spk, name2text, name2wavpath)

    @property
    def info(self):
        return self.info_dict




if __name__=="__main__":
    pass
    # raw_dir = "/data/duhu/dbase/zh_cn/readCorpus/King-ASR-027"
    # kformat_dir = os.path.join("data", os.path.basename(raw_dir))
    # inputGenerator = InputDialogCorpusFormat(raw_dir, kformat_dir=kformat_dir, item_index=0)
    # inputGenerator.process()


    # King-ASR-052 King-ASR-042 King-ASR-023 King-ASR-059 King-ASR-065 King-ASR-084
    # raw_dir = f"/data/duhu/dbase/zh_cn/readCorpus/King-ASR-084" 
    # kformat_dir = os.path.join("data", os.path.basename(raw_dir))
    # inputGenerator = InputReadCorpusFormat(raw_dir, kformat_dir=kformat_dir, item_index=1)
    # inputGenerator.process()

    # King-ASR-113 一部分是item 0 一部分是 item 1

    
    # raw_dir = f"/data/duhu/dbase/zh_cn/trdata_new/TR2020064" 
    # kformat_dir = os.path.join("data", os.path.basename(raw_dir))
    # inputGenerator = InputVIPKIDTR2020064(raw_dir, kformat_dir=kformat_dir)
    # inputGenerator.process()

    # raw_dir = f"/data/duhu/dbase/ja-JP" 
    # kformat_dir = "data"
    # inputGenerator = InputBytedanceJAJP(raw_dir, kformat_dir=kformat_dir)
    # inputGenerator.process()


    # raw_dir = f"/data/duhu/dbase/ko-KR/bytedance" 
    # kformat_dir = "data"
    # inputGenerator = InputBytedanceKOKR(raw_dir, kformat_dir=kformat_dir)
    # inputGenerator.process()

    # raw_dir = f"/data/duhu/dbase/pt-BR" 
    # kformat_dir = "data"
    # inputGenerator = InputBytedancePTBR(raw_dir, kformat_dir=kformat_dir)
    # inputGenerator.process()
